"""
Database models for the finance automation system
"""
from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime, Boolean, Text, ForeignKey, Index, UniqueConstraint, Date
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, relationship
from datetime import datetime
import os
import base64
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC

Base = declarative_base()


def _get_cipher():
    """
    Derive an encryption key from PLAID_SECRET so tokens are encrypted at rest.
    No separate key to manage — tied to your existing credential.
    """
    secret = os.getenv('PLAID_SECRET', 'fallback-dev-secret-change-in-production')
    salt = b'finance_automation_salt_v1'  # Fixed salt is fine here; secret is the entropy
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=100_000)
    key = base64.urlsafe_b64encode(kdf.derive(secret.encode()))
    return Fernet(key)


def encrypt_token(token: str) -> str:
    """Encrypt a Plaid access token for storage"""
    return _get_cipher().encrypt(token.encode()).decode()


def decrypt_token(encrypted: str) -> str:
    """Decrypt a stored Plaid access token"""
    return _get_cipher().decrypt(encrypted.encode()).decode()


class PlaidItem(Base):
    """
    Represents a Plaid Item (one per bank connection).
    Stores the access token encrypted so it survives server restarts.
    """
    __tablename__ = 'plaid_items'

    id = Column(Integer, primary_key=True)
    item_id = Column(String(100), unique=True, nullable=False, index=True)
    institution_name = Column(String(200))          # e.g. "Chase" — user-renameable
    institution_id = Column(String(50))             # Plaid's immutable ID e.g. "ins_3" — used for matching
    access_token_enc = Column(Text, nullable=False) # AES-encrypted via Fernet
    cursor = Column(Text, nullable=True)            # Plaid sync cursor — persisted here
    last_synced_at = Column(DateTime, nullable=True)
    last_error_code = Column(String(100), nullable=True)    # e.g. ITEM_LOGIN_REQUIRED
    last_error_message = Column(Text, nullable=True)        # Human-readable error detail
    last_error_at = Column(DateTime, nullable=True)         # When the error was first recorded
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    @property
    def access_token(self) -> str:
        """Decrypt and return the access token"""
        return decrypt_token(self.access_token_enc)

    @access_token.setter
    def access_token(self, raw_token: str):
        """Encrypt and store the access token"""
        self.access_token_enc = encrypt_token(raw_token)


class Account(Base):
    """Bank/credit card accounts linked via Plaid"""
    __tablename__ = 'accounts'
    
    id = Column(Integer, primary_key=True)
    plaid_account_id = Column(String(100), unique=True, nullable=True, index=True)  # NULL for manual accounts
    persistent_account_id = Column(String(200), nullable=True, index=True)  # Stable across re-links (Plaid persistent_account_id); no UNIQUE constraint (edge-case recovery may have duplicates temporarily)
    plaid_item_id = Column(String(100), nullable=True, index=True)  # FK to plaid_items.item_id; NULL for manual
    institution_id = Column(String(50), nullable=True, index=True)  # Copied from PlaidItem; persists after sever-plaid so re-link matching stays institution-scoped
    # Stable 12-char identity hash: SHA256(institution_id|mask|account_type)[:12]
    # Written at exchange-token time; used as primary re-link matching key.
    # Survives sever-plaid because it's stored on the account row itself.
    account_hash = Column(String(16), nullable=True, index=True)
    account_name = Column(String(100), nullable=False)  # e.g., "Chase 8997"
    account_type = Column(String(50))  # checking, credit, etc.
    official_name = Column(String(200))
    mask = Column(String(10))  # Last 4 digits
    is_manual = Column(Boolean, default=False)
    starting_balance = Column(Float, default=0)       # Balance when tracking began
    start_date = Column(DateTime, nullable=True)       # Date starting_balance applies to
    notes = Column(Text, nullable=True)                # Optional user notes
    product_id = Column(Integer, ForeignKey('card_products.id'), nullable=True, index=True)  # Links account directly to card product
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.utcnow)

    # Plaid Liabilities product — populated by POST /api/plaid/sync-liabilities
    # Applies to credit cards and loan accounts; NULL for depository/investment accounts.
    liability_min_payment     = Column(Float,    nullable=True)   # Minimum amount due this cycle
    liability_next_due_date   = Column(DateTime, nullable=True)   # When next payment is due
    liability_last_statement_bal = Column(Float, nullable=True)   # Balance as of last statement
    liability_last_payment    = Column(Float,    nullable=True)   # Last payment amount
    liability_last_payment_date = Column(DateTime, nullable=True) # Date of last payment
    liability_purchase_apr    = Column(Float,    nullable=True)   # Purchase APR % (credit cards)

    transactions = relationship("Transaction", back_populates="account")
    card = relationship("Card", back_populates="account", uselist=False, foreign_keys="Card.account_id")
    product = relationship("CardProduct")


class Transaction(Base):
    """Individual transactions from Plaid"""
    __tablename__ = 'transactions'
    
    id = Column(Integer, primary_key=True)
    plaid_transaction_id = Column(String(100), unique=True, nullable=True, index=True)  # NULL for manual txns
    account_id = Column(Integer, ForeignKey('accounts.id'), nullable=False, index=True)
    
    # Core transaction data
    date = Column(DateTime, nullable=False, index=True)
    amount = Column(Float, nullable=False)  # Negative for expenses, positive for income
    description_raw = Column(String(500), nullable=False)
    description_clean = Column(String(500))  # Cleaned/standardized description
    merchant_name = Column(String(200))
    
    # Categorization
    action = Column(String(50), index=True)  # Income, Expense, Transfer, Depreciation
    category_auto = Column(String(100))  # Auto-assigned category
    category_manual = Column(String(100))  # User-corrected category
    category_confidence = Column(Float)  # Confidence score 0-1
    
    # Status
    needs_review = Column(Boolean, default=True, index=True)
    is_locked = Column(Boolean, default=False, index=True)  # Protects manual edits from sync overwrites
    is_split = Column(Boolean, default=False)
    parent_transaction_id = Column(Integer, ForeignKey('transactions.id'), nullable=True)

    # Enrichment source — how this transaction got its category/description
    # Values: 'rule' (matched a categorization rule), 'llm' (Groq LLM),
    #         'fallback' (LLM unavailable), 'manual' (user-created), None (not yet enriched)
    enrichment_source = Column(String(20), nullable=True)

    # Import tracking — for CSV/OFX imported transactions (NULL for Plaid)
    # import_hash: SHA-256 of (account_id + date + amount + description + occurrence_index)
    #              used for deduplication on re-import; unique constraint allows NULL (Plaid txns)
    import_hash = Column(String(64), unique=True, nullable=True, index=True)
    # import_source: 'plaid' | 'csv' | 'ofx' | 'manual'
    import_source = Column(String(20), nullable=True)

    # Content hash — stable identity that survives Plaid re-links.
    # Format: SHA256(account_id|date|amount|description_raw)[:14] + "-" + NN (zero-padded suffix)
    # e.g. "a1b2c3d4e5f6a7-00". The suffix differentiates identical transactions
    # on the same day (same merchant, same amount). No UNIQUE constraint — the
    # base+suffix pair is unique by construction at insert time.
    content_hash = Column(String(20), nullable=True, index=True)

    # GCB (Gift Card Business) tagging
    gcb_tagged = Column(Boolean, default=False, index=True)  # Legacy — kept for backward compat
    is_gcb = Column(Boolean, default=False, index=True)       # New canonical GCB flag (Section 3b)

    # Points tracking
    points_category = Column(String(100), nullable=True)   # From PointsCategory table
    card_id = Column(Integer, ForeignKey('cards.id'), nullable=True)  # Which card was used

    # Loan payment tracking — set when this transaction is linked to a loan payment
    loan_id = Column(Integer, ForeignKey('loans.id'), nullable=True, index=True)

    # Exclude flag — user can exclude a transaction from all totals/balances without deleting it
    # Useful for pending transactions that already entered the DB and should be ignored.
    is_excluded = Column(Boolean, default=False, index=True)

    # Manual override for the signed points-earn value (see compute_points_earn() in
    # main.py) — sticks even if the auto-classification logic later changes.
    points_earn_override = Column(Float, nullable=True)

    # Free-text tag for who made this purchase (e.g. "Omer", "Daniella") — manual only,
    # Plaid gives no cardholder-level signal on shared/authorized-user accounts.
    # Lets a SpendChallenge.spender_filter scope its spend calc to one person's purchases.
    spender = Column(String(100), nullable=True)

    # Metadata
    year = Column(Integer, index=True)
    month = Column(Integer, index=True)
    day = Column(Integer)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    reviewed_at = Column(DateTime, nullable=True)
    
    # Relationships
    account = relationship("Account", back_populates="transactions")
    splits = relationship("Transaction", remote_side=[id], foreign_keys=[parent_transaction_id])
    
    # Composite indexes for common queries
    __table_args__ = (
        Index('ix_transactions_date_account', 'date', 'account_id'),
        Index('ix_transactions_needs_review_date', 'needs_review', 'date'),
        Index('ix_transactions_year_month', 'year', 'month'),
    )
    
    @property
    def category_final(self):
        """Return manual category if set, otherwise auto category"""
        return self.category_manual or self.category_auto
    
    @property
    def amount_final(self):
        """Return actual transaction amount"""
        return self.amount


class CategorizationRule(Base):
    """Rules for automatic categorization based on description patterns"""
    __tablename__ = 'categorization_rules'

    id = Column(Integer, primary_key=True)
    priority = Column(Integer, default=100, index=True)  # Lower = higher priority
    priority_order = Column(Integer, default=0)  # Sub-priority within same priority level (Section 2D)
    match_type = Column(String(50), nullable=False)  # contains, equals, starts_with, regex
    pattern = Column(String(500), nullable=False, index=True)

    # Actions
    set_action = Column(String(50))  # Income, Expense, Transfer, etc.
    set_category = Column(String(100))
    set_description = Column(String(500))  # Standardized description
    clean_description = Column(String(500))  # Cleaned/normalized output description (Section 2D)

    # Metadata
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow)
    notes = Column(Text)

    # Learning stats
    times_matched = Column(Integer, default=0)
    times_accepted = Column(Integer, default=0)
    times_rejected = Column(Integer, default=0)
    
    @property
    def accuracy(self):
        """Calculate rule accuracy"""
        total = self.times_accepted + self.times_rejected
        if total == 0:
            return None
        return self.times_accepted / total


class Category(Base):
    """Master list of expense/income categories"""
    __tablename__ = 'categories'
    
    id = Column(Integer, primary_key=True)
    name = Column(String(100), unique=True, nullable=False, index=True)
    parent_category = Column(String(100), nullable=True)  # For future subcategories
    category_type = Column(String(50), index=True)  # expense, income, or both
    is_active = Column(Boolean, default=True)
    display_order = Column(Integer, default=100)
    created_at = Column(DateTime, default=datetime.utcnow)


class UserCorrection(Base):
    """Track user corrections to improve ML over time"""
    __tablename__ = 'user_corrections'
    
    id = Column(Integer, primary_key=True)
    transaction_id = Column(Integer, ForeignKey('transactions.id'), nullable=False, index=True)
    
    # What changed
    old_category = Column(String(100))
    new_category = Column(String(100), nullable=False)
    old_action = Column(String(50))
    new_action = Column(String(50))
    
    # Context for learning
    description = Column(String(500))
    merchant_name = Column(String(200))
    amount = Column(Float)
    
    created_at = Column(DateTime, default=datetime.utcnow)



class TransactionSplit(Base):
    """
    Split line items for a parent transaction (Section 3a).
    Sum of split amounts must equal parent transaction amount.
    """
    __tablename__ = 'transaction_splits'

    id = Column(Integer, primary_key=True)
    parent_transaction_id = Column(Integer, ForeignKey('transactions.id'), nullable=False, index=True)
    amount = Column(Float, nullable=False)
    description = Column(String(500))
    category = Column(String(100))
    action = Column(String(50), nullable=True)  # Type per split line (Section 4F)
    is_gcb = Column(Boolean, default=False)  # GCB tag per split
    notes = Column(Text, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)


class Loan(Base):
    """
    Loan tracking — mortgages, auto loans, student loans, etc. (Section 1).
    Links to an account for balance tracking.
    """
    __tablename__ = 'loans'

    id = Column(Integer, primary_key=True)
    account_id = Column(Integer, ForeignKey('accounts.id'), nullable=True)  # Linked liability account
    lender = Column(String(200), nullable=False)
    loan_type = Column(String(50), nullable=False)  # mortgage, auto, student, personal, other
    original_principal = Column(Float, nullable=False)
    current_balance = Column(Float, nullable=True)   # Known balance as of balance_date
    balance_date = Column(DateTime, nullable=True)   # Date when current_balance was recorded
    remaining_term_months = Column(Integer, nullable=True)  # Remaining months as of balance_date
    interest_rate = Column(Float, nullable=True)     # Annual rate as percentage (e.g. 6.5)
    term_months = Column(Integer, nullable=True)     # Original total term in months
    monthly_payment = Column(Float, nullable=True)   # Total monthly payment (PITI)
    property_tax_monthly = Column(Float, nullable=True)   # Escrow: monthly property tax portion
    insurance_monthly = Column(Float, nullable=True)      # Escrow: monthly insurance portion
    payment_account_id = Column(Integer, ForeignKey('accounts.id'), nullable=True)  # Checking account that pays
    payment_due_day = Column(Integer, nullable=True)      # Day of month payment is due (1-31)
    start_date = Column(DateTime, nullable=True)
    maturity_date = Column(DateTime, nullable=True)
    is_active = Column(Boolean, default=True)
    notes = Column(Text, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class AccountMonthlySnapshot(Base):
    """
    Monthly opening/closing balance snapshots per account.
    Populated by the Balance Sync tool and updated after every Plaid transaction sync.
    Formula: opening_balance + SUM(transactions in month) = closing_balance
    """
    __tablename__ = 'account_monthly_snapshots'

    id              = Column(Integer, primary_key=True)
    account_id      = Column(Integer, ForeignKey('accounts.id', ondelete='CASCADE'), nullable=False, index=True)
    year            = Column(Integer, nullable=False)
    month           = Column(Integer, nullable=False)   # 1–12
    opening_balance = Column(Float, nullable=False)
    closing_balance = Column(Float, nullable=False)
    synced_at       = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    __table_args__ = (
        UniqueConstraint('account_id', 'year', 'month', name='uq_account_month'),
    )


class BalanceObservation(Base):
    """
    Plaid-reported balance snapshots captured every transaction sync.
    Used as self-correcting anchors for daily balance calculations.
    Each row records: what Plaid says, what we computed, and the drift.
    """
    __tablename__ = 'balance_observations'

    id               = Column(Integer, primary_key=True)
    account_id       = Column(Integer, ForeignKey('accounts.id', ondelete='CASCADE'), nullable=False)
    observed_at      = Column(DateTime, nullable=False, default=datetime.utcnow)
    plaid_balance    = Column(Float, nullable=False)       # Signed (credit/loan negated)
    computed_balance = Column(Float, nullable=True)        # Transaction-derived balance at observation time
    delta            = Column(Float, nullable=True)        # plaid_balance - computed_balance
    source           = Column(String(20), default='sync')  # 'sync', 'balance_sync', 'manual'

    __table_args__ = (
        Index('ix_balance_obs_account_observed', 'account_id', 'observed_at'),
    )


class BudgetTarget(Base):
    """
    Monthly budget targets per category (Section 4).
    Unique on (year, month, category) — one amount per cell.
    """
    __tablename__ = 'budget_targets'

    id = Column(Integer, primary_key=True)
    year = Column(Integer, nullable=False, index=True)
    month = Column(Integer, nullable=False, index=True)  # 1-12
    category = Column(String(100), nullable=False, index=True)
    amount = Column(Float, nullable=False, default=0)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    __table_args__ = (
        Index('ix_budget_targets_ymc', 'year', 'month', 'category', unique=True),
    )


class PointsCategory(Base):
    """Universal points/miles earning categories based on card issuer classification.

    Two-level hierarchy for points calculation:
      L1 (broad)  — Airlines, Hotels, Dining, etc.  parent_key = None
      L2 (brand)  — United, Hilton, Best Buy, etc.  parent_key = L1 name

    Earn-rate waterfall: try L2 rate first, fall back to L1, fall back to base.
    """
    __tablename__ = 'points_categories'

    id = Column(Integer, primary_key=True)
    name = Column(String(100), unique=True, nullable=False)
    display_order = Column(Integer, default=100)
    is_active = Column(Boolean, default=True)
    # L1/L2 hierarchy — stores the parent category's name (e.g. "Airlines" for "United")
    parent_key = Column(String(100), nullable=True, index=True)

    merchant_mappings = relationship("MerchantPointsMapping", back_populates="points_category")


class Card(Base):
    """Credit/debit cards — one per physical card"""
    __tablename__ = 'cards'

    id = Column(Integer, primary_key=True)
    card_id = Column(String(50), unique=True, nullable=False, index=True)  # e.g. "VISA 1677"
    last_four = Column(Integer, nullable=True)      # Actual last 4 of card number
    issuer = Column(String(50))                     # CHASE, AMEX, CITI, BOA, etc.
    brand = Column(String(100))                     # Hyatt, Hilton, Chase, etc.
    card_name = Column(String(200))                 # Sapphire Preferred, Freedom, etc.
    network = Column(String(20))                    # VISA, AMEX, MC, DISCOVER
    issue_date = Column(DateTime, nullable=True)
    close_date = Column(DateTime, nullable=True)    # Date card was closed (null = active)
    annual_fee = Column(Float, nullable=True)
    credit_limit = Column(Float, nullable=True)
    statement_close_day = Column(Integer, nullable=True)   # Day of month (1-31)
    payment_due_day = Column(Integer, nullable=True)       # Day of month (1-31)
    plaid_account_id = Column(String(100), nullable=True)  # Legacy string link — prefer account_id
    account_id = Column(Integer, ForeignKey('accounts.id'), nullable=True, index=True)  # Proper FK to Account
    payment_account_id = Column(Integer, ForeignKey('accounts.id'), nullable=True)  # Checking account that pays this card
    product_id = Column(Integer, ForeignKey('card_products.id'), nullable=True, index=True)  # Links to card product
    ecosystem_id = Column(Integer, ForeignKey('points_ecosystems.id'), nullable=True)  # Legacy — prefer product.ecosystem
    is_active = Column(Boolean, default=True)
    notes = Column(Text, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    account = relationship("Account", back_populates="card", foreign_keys=[account_id])
    payment_account = relationship("Account", foreign_keys=[payment_account_id])
    product = relationship("CardProduct", back_populates="cards")
    ecosystem_rel = relationship("PointsEcosystem", back_populates="cards")
    merchant_mappings = relationship("MerchantPointsMapping", back_populates="card")
    spend_challenges = relationship('SpendChallenge', back_populates='card', cascade='all, delete-orphan')


class PointsEcosystem(Base):
    """Points/miles currency ecosystem with valuations"""
    __tablename__ = 'points_ecosystems'

    id = Column(Integer, primary_key=True)
    name = Column(String(50), unique=True, nullable=False)    # "Chase UR", "Amex MR", "Hilton", etc.
    currency_name = Column(String(100))                        # "Ultimate Rewards", "Membership Rewards"
    eco_type = Column(String(20))                              # "Flexible", "Hotel", "Airline", "Cash"
    conservative_cpp = Column(Float, default=1.0)              # Conservative cents per point
    your_cpp = Column(Float, default=1.0)                      # User's personal valuation
    is_cash_back = Column(Boolean, default=False)              # True = flat cash back (always 1 cpp)
    conservative_basis = Column(String(200), nullable=True)    # Explanation of conservative value

    products = relationship("CardProduct", back_populates="ecosystem_rel")
    # Legacy: cards may still reference ecosystem_id directly
    cards = relationship("Card", back_populates="ecosystem_rel")


class CardProduct(Base):
    """
    One row per card PRODUCT (e.g., 'Chase Sapphire Preferred').
    Multiple physical Cards may reference the same product
    (Omer's CSP + Daniella's CSP = 2 cards, 1 product).
    Earning rates and benefits live here, not on Card.
    """
    __tablename__ = 'card_products'

    id = Column(Integer, primary_key=True)
    product_key = Column(String(50), unique=True, nullable=False, index=True)  # e.g., 'chase_sapphire_preferred'
    card_name = Column(String(200))                                             # Display name
    ecosystem_id = Column(Integer, ForeignKey('points_ecosystems.id'), nullable=True)
    status = Column(String(20), default='active')                               # active, closed, planned
    notes = Column(Text, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    ecosystem_rel = relationship("PointsEcosystem", back_populates="products")
    rewards = relationship("CardProductReward", back_populates="product", cascade="all, delete-orphan")
    benefits = relationship("CardBenefit", back_populates="product", cascade="all, delete-orphan")
    cards = relationship("Card", back_populates="product")


class CardProductReward(Base):
    """
    Category multipliers for a card product.
    Base rate: points_category_id = NULL, is_base_rate = True.
    Category bonus: ADDITIONAL points on top of base.
    Total for category = base + category bonus.
    """
    __tablename__ = 'card_product_rewards'

    id = Column(Integer, primary_key=True)
    product_id = Column(Integer, ForeignKey('card_products.id', ondelete='CASCADE'), nullable=False)
    points_category_id = Column(Integer, ForeignKey('points_categories.id', ondelete='CASCADE'), nullable=True)
    multiplier = Column(Float, nullable=False)     # Points per dollar (base or additional)
    is_base_rate = Column(Boolean, default=False)  # True = applies to ALL spend
    reward_type = Column(String, nullable=False, server_default='fixed', default='fixed')
    # reward_type values:
    #   'fixed'             — standard fixed-rate bonus (default)
    #   'auto_top_category' — dynamic best-category (e.g. Citi Custom Cash 5% on top eligible cat)

    product = relationship("CardProduct", back_populates="rewards")
    points_category = relationship("PointsCategory")

    __table_args__ = (
        Index('ix_product_reward_prod_cat', 'product_id', 'points_category_id'),
    )


class CardBenefit(Base):
    """
    Recurring credits/perks for a card product.
    e.g., Amex Platinum $200 airline credit, Hilton Aspire $250 resort credit.
    """
    __tablename__ = 'card_benefits'

    id = Column(Integer, primary_key=True)
    product_id = Column(Integer, ForeignKey('card_products.id', ondelete='CASCADE'), nullable=False)
    benefit_name = Column(String(200), nullable=False)        # "Airline Credit", "Resort Credit"
    amount = Column(Float, nullable=False)                     # Dollar value per cycle
    reset_frequency = Column(String(20), default='annual')     # "annual", "semi-annual", "monthly", "calendar_year"
    trigger_category = Column(String(100), nullable=True)      # Spending category that triggers the credit (if any)
    notes = Column(Text, nullable=True)
    tracking_type = Column(String(20), default='periodic')     # "periodic" (usage tracker) or "by_use" (no cadence)

    product = relationship("CardProduct", back_populates="benefits")
    usage = relationship("BenefitUsage", back_populates="benefit", cascade="all, delete-orphan")


class BenefitUsage(Base):
    """Tracks per-cycle usage of a card benefit."""
    __tablename__ = 'benefit_usage'

    id = Column(Integer, primary_key=True)
    benefit_id = Column(Integer, ForeignKey('card_benefits.id', ondelete='CASCADE'), nullable=False)
    card_id = Column(Integer, ForeignKey('cards.id', ondelete='CASCADE'), nullable=True)  # Which physical card used it
    cycle = Column(String(20), nullable=False)                 # "2026", "2026-H1", "2026-03" depending on frequency
    amount_used = Column(Float, default=0)                     # How much of the benefit has been used
    confirmed = Column(Boolean, default=False)                 # User-confirmed usage
    notes = Column(Text, nullable=True)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    benefit = relationship("CardBenefit", back_populates="usage")

    __table_args__ = (
        Index('ix_benefit_usage_benefit_cycle', 'benefit_id', 'cycle'),
    )


class SpendChallenge(Base):
    """Time-boxed earning challenge attached to a specific card instance.

    challenge_type values:
      'rate_cap'          — earn bonus_amount pts/$ on all spend, up to spend_cap
      'threshold_bonus'   — earn bonus_amount pts/$ on all spend IF >= spend_threshold
      'category_rate_cap' — like rate_cap but only qualifying CSC spend counts
      'sub'               — flat bonus_amount pts IF spend >= spend_threshold (Sign-Up Bonus)
      'annual_threshold'  — flat benefit (e.g. free night cert) IF spend >= threshold; resets Jan 1

    bonus_type values: 'per_dollar' | 'flat' | 'benefit'
      'benefit' = non-points reward (free night cert, status, etc.) — bonus_amount = # of rewards
    """
    __tablename__ = 'spend_challenges'

    id              = Column(Integer, primary_key=True)
    card_id         = Column(Integer, ForeignKey('cards.id', ondelete='CASCADE'), nullable=False, index=True)
    name            = Column(String(200), nullable=False)
    challenge_type  = Column(String(30), nullable=False)

    start_date      = Column(Date, nullable=False)
    end_date        = Column(Date, nullable=False)
    # Effective spend tracking starts at max(start_date, activation_date).
    # Use when the card was opened after the challenge period started (e.g. SUB clock).
    activation_date = Column(Date, nullable=True)

    # What you earn
    bonus_type      = Column(String(20), nullable=False)   # 'per_dollar' | 'flat' | 'benefit'
    bonus_amount    = Column(Float, nullable=False)        # pts/$ | flat pts | # of rewards

    # Conditions
    spend_cap       = Column(Float, nullable=True)   # max eligible spend (rate_cap, category_rate_cap)
    spend_threshold = Column(Float, nullable=True)   # min spend to unlock (threshold_bonus, sub, annual_threshold)
    spender_filter  = Column(String(100), nullable=True)  # match Transaction.spender; NULL = anyone's spend counts

    # Cached progress — recalculated from transactions on demand
    current_spend   = Column(Float, default=0)
    bonus_unlocked  = Column(Boolean, default=False)

    is_active       = Column(Boolean, default=True)
    notes           = Column(Text, nullable=True)
    created_at      = Column(DateTime, default=datetime.utcnow)

    card             = relationship('Card', back_populates='spend_challenges')
    # Direct 1-to-many relationships to the junction tables.
    # Using simple FK relationships (not complex secondary joins) for reliability.
    # Access card_ids via:  [lnk.card_id      for lnk in challenge.card_links]
    # Access cat names via: [lnk.category_name for lnk in challenge.category_links]
    card_links     = relationship('ChallengeCardLink',     cascade='all, delete-orphan',
                                  foreign_keys='ChallengeCardLink.challenge_id')
    category_links = relationship('ChallengeCategoryLink', cascade='all, delete-orphan',
                                  foreign_keys='ChallengeCategoryLink.challenge_id')


class ChallengeCardLink(Base):
    """Junction table: SpendChallenge ↔ additional Card (many-to-many)."""
    __tablename__ = 'challenge_card_links'
    challenge_id = Column(Integer, ForeignKey('spend_challenges.id', ondelete='CASCADE'), primary_key=True)
    card_id      = Column(Integer, ForeignKey('cards.id',            ondelete='CASCADE'), primary_key=True)


class ChallengeCategoryLink(Base):
    """Junction table: SpendChallenge ↔ PointsCategory (multi-category challenges).
    No FK on category_name — stored as plain string to avoid PG UNIQUE constraint dependency."""
    __tablename__ = 'challenge_category_links'
    challenge_id  = Column(Integer, ForeignKey('spend_challenges.id', ondelete='CASCADE'), primary_key=True)
    category_name = Column(String(100), nullable=False, primary_key=True)


class Redemption(Base):
    """Pure value capture for points actually redeemed — e.g. Hilton Honors
    points spent on a hotel stay. Deliberately carries no transfer/source
    info: once points land in an ecosystem they're fungible, so there's no
    clean way to attribute redemption value back through a prior transfer.
    See Transfer below for the (value-neutral) point-movement side. Lets you
    compare realized cpp (cash_value_usd / points_redeemed) against an
    ecosystem's assumed your_cpp."""
    __tablename__ = 'redemptions'

    id                 = Column(Integer, primary_key=True)
    ecosystem_id       = Column(Integer, ForeignKey('points_ecosystems.id'), nullable=False, index=True)
    points_redeemed    = Column(Float, nullable=False)
    redemption_date    = Column(Date, nullable=False)
    description        = Column(String(300), nullable=False)
    cash_value_usd     = Column(Float, nullable=False)
    notes              = Column(Text, nullable=True)
    created_at         = Column(DateTime, default=datetime.utcnow)

    ecosystem = relationship('PointsEcosystem', foreign_keys=[ecosystem_id])


class TransferRatio(Base):
    """Current (and historical) transfer ratio between an ecosystem pair,
    e.g. Amex MR -> Hilton Honors = 2.0 (destination points per 1 source
    point). Effective-dated: editing a ratio closes the old row's
    effective_to and opens a new one rather than overwriting it, so past
    Transfers (which snapshot their own base_ratio_used) stay accurate
    regardless of later ratio changes. effective_to = NULL means "current"."""
    __tablename__ = 'transfer_ratios'

    id                        = Column(Integer, primary_key=True)
    source_ecosystem_id      = Column(Integer, ForeignKey('points_ecosystems.id'), nullable=False, index=True)
    destination_ecosystem_id = Column(Integer, ForeignKey('points_ecosystems.id'), nullable=False, index=True)
    base_ratio                = Column(Float, nullable=False)
    effective_from            = Column(Date, nullable=False)
    effective_to              = Column(Date, nullable=True)
    created_at                = Column(DateTime, default=datetime.utcnow)

    source_ecosystem      = relationship('PointsEcosystem', foreign_keys=[source_ecosystem_id])
    destination_ecosystem = relationship('PointsEcosystem', foreign_keys=[destination_ecosystem_id])


class Transfer(Base):
    """A value-neutral point-movement event between two ecosystems (e.g.
    100,000 Amex MR -> 140,000 Marriott Bonvoy with a 40% transfer bonus).
    Self-contained/immutable: base_ratio_used and points_received are
    snapshotted at creation time (typically defaulted from the current
    TransferRatio for the pair, but editable per-transfer), so this row
    never needs to be recomputed if TransferRatio changes later."""
    __tablename__ = 'transfers'

    id                        = Column(Integer, primary_key=True)
    source_ecosystem_id      = Column(Integer, ForeignKey('points_ecosystems.id'), nullable=False, index=True)
    destination_ecosystem_id = Column(Integer, ForeignKey('points_ecosystems.id'), nullable=False, index=True)
    points_sent               = Column(Float, nullable=False)
    base_ratio_used           = Column(Float, nullable=False)
    bonus_pct                 = Column(Float, nullable=True)  # e.g. 0.40 for a 40% promo bonus
    points_received           = Column(Float, nullable=False)
    transfer_date              = Column(Date, nullable=False)
    notes                      = Column(Text, nullable=True)
    created_at                  = Column(DateTime, default=datetime.utcnow)

    source_ecosystem      = relationship('PointsEcosystem', foreign_keys=[source_ecosystem_id])
    destination_ecosystem = relationship('PointsEcosystem', foreign_keys=[destination_ecosystem_id])


class PointsBalanceSnapshot(Base):
    """Manual balance checkpoint for an ecosystem — 'I logged into my Amex
    account and it says 40,320 points as of today.' Corrects for drift
    between the computed running balance (all-time earned minus
    redeemed/transferred) and reality — unlogged promo bonuses, benefit
    credits, redemptions made outside this app, etc. The most recent
    snapshot becomes the baseline for current_balance; activity before its
    date is ignored (assumed already folded into the snapshotted value)."""
    __tablename__ = 'points_balance_snapshots'

    id            = Column(Integer, primary_key=True)
    ecosystem_id  = Column(Integer, ForeignKey('points_ecosystems.id'), nullable=False, index=True)
    balance       = Column(Float, nullable=False)
    snapshot_date = Column(Date, nullable=False)
    notes         = Column(Text, nullable=True)
    created_at    = Column(DateTime, default=datetime.utcnow)

    ecosystem = relationship('PointsEcosystem', foreign_keys=[ecosystem_id])


# Legacy alias — kept for backward compatibility during migration
CardEarningRate = CardProductReward


class MerchantPointsMapping(Base):
    """Maps merchant + card to points category for points calculation"""
    __tablename__ = 'merchant_points_mappings'

    id = Column(Integer, primary_key=True)
    merchant_pattern = Column(String(200), nullable=False, index=True)  # Merchant name/pattern
    card_id = Column(Integer, ForeignKey('cards.id'), nullable=True)    # Null = applies to all cards
    points_category_id = Column(Integer, ForeignKey('points_categories.id'), nullable=False)
    notes = Column(Text, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)

    card = relationship("Card", back_populates="merchant_mappings")
    points_category = relationship("PointsCategory", back_populates="merchant_mappings")


class DuplicateIgnore(Base):
    """
    Stores pairs of account IDs that the user has confirmed are NOT duplicates
    (e.g. two genuinely different Amex cards that happen to share the same mask).
    The scan will never flag these two accounts together again.
    Always store with account_id_a < account_id_b for uniqueness.
    """
    __tablename__ = 'duplicate_ignore'

    id = Column(Integer, primary_key=True)
    account_id_a = Column(Integer, nullable=False)   # lower of the two IDs
    account_id_b = Column(Integer, nullable=False)   # higher of the two IDs
    created_at = Column(DateTime, default=datetime.utcnow)

    __table_args__ = (
        UniqueConstraint('account_id_a', 'account_id_b', name='uq_dup_ignore_pair'),
    )


class CashFlowOverlay(Base):
    """
    Explicit upcoming cash flows displayed below the Daily Balances grid.
    Sources:
      'manual'       — user-created entry (paycheck, rent, etc.)
      'cc_payment'   — auto-generated from card due-date + balance
      'loan_payment' — auto-generated from loan monthly payment
    amount is signed: negative = outflow (payment), positive = inflow (paycheck).
    """
    __tablename__ = 'cash_flow_overlays'

    id             = Column(Integer, primary_key=True)
    description    = Column(String(200), nullable=False)
    amount         = Column(Float, nullable=False)          # + inflow / – outflow
    flow_date      = Column(Date, nullable=False, index=True)
    source         = Column(String(20), default='manual')   # manual|cc_payment|loan_payment
    account_id     = Column(Integer, ForeignKey('accounts.id'), nullable=True)
    is_recurring   = Column(Boolean, default=False)
    recurrence_day = Column(Integer, nullable=True)         # day of month (1–31) if recurring
    is_active      = Column(Boolean, default=True)
    created_at     = Column(DateTime, default=datetime.utcnow)
    updated_at     = Column(DateTime, default=datetime.utcnow)

    account = relationship('Account', foreign_keys=[account_id])


class PlannedPurchase(Base):
    """
    One-off planned large purchases for liquidity forecasting.
    """
    __tablename__ = 'planned_purchases'

    id            = Column(Integer, primary_key=True)
    name          = Column(String(255), nullable=False)
    amount        = Column(Float, nullable=False)          # Always positive outflow
    expected_date = Column(Date, nullable=False)
    vendor_tag    = Column(String(100), nullable=True)
    status        = Column(String(50), default='pending')  # pending|completed|cancelled
    created_at    = Column(DateTime, default=datetime.utcnow)
    updated_at    = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class SalaryPayment(Base):
    """
    One income event (paycheck, HSA/FSA contribution, etc.) for one person on one date.
    Each SalaryPayment has one or more SalaryAllocation rows that specify how much
    is deposited into each account.  Future payments are projected in the daily-balance
    grid via get_daily_balances.
    """
    __tablename__ = 'salary_payments'

    id           = Column(Integer, primary_key=True)
    payment_date = Column(Date, nullable=False, index=True)
    description  = Column(String(200), nullable=False)   # "Salary", "HSA Contribution", …
    person       = Column(String(100), nullable=False)   # "Omer", "Daniella", …
    is_active    = Column(Boolean, default=True)
    created_at   = Column(DateTime, default=datetime.utcnow)

    allocations = relationship(
        'SalaryAllocation', back_populates='salary_payment',
        cascade='all, delete-orphan',
    )


class SalaryAllocation(Base):
    """Per-account deposit amount for a SalaryPayment (always a positive inflow)."""
    __tablename__ = 'salary_allocations'

    id                = Column(Integer, primary_key=True)
    salary_payment_id = Column(Integer, ForeignKey('salary_payments.id'), nullable=False)
    account_id        = Column(Integer, ForeignKey('accounts.id'), nullable=False)
    amount            = Column(Float, nullable=False)    # always positive

    salary_payment = relationship('SalaryPayment', back_populates='allocations')
    account        = relationship('Account', foreign_keys=[account_id])


# ---------------------------------------------------------------------------
# Category clean-up remap — old name → new canonical name.
# Applied once per startup; idempotent (re-running is safe).
# Covers: transaction category_auto/manual, budget_targets, rules.
# ---------------------------------------------------------------------------
_CAT_REMAP = [
    # Utilities consolidation
    ('Phone',             'Utilities'),
    ('Internet',          'Utilities'),
    ('Water',             'Utilities'),
    ('Electricity',       'Utilities'),
    # Entertainment consolidation
    ('Leisure',           'Entertainment'),
    ('Books',             'Entertainment'),
    ('Events',            'Entertainment'),
    # Education consolidation
    ('Music Lessons',     'Education'),
    ('Tutoring',          'Education'),
    ('Studies',           'Education'),
    # Business
    ('Consulting',        'Business'),
    # Investment / interest
    ('Investments',       'Investment Gain (Loss)'),
    ('Investment Income', 'Investment Gain (Loss)'),
    ('Interest Income',   'Investment Gain (Loss)'),
    # Family support
    ('Siblings',          'For Others'),
    ('Parents',           'For Others'),
    # Misc
    ('Lottery',           'Other'),
    ('Dry Cleaning',      'Self Care'),
    # Name standardisation
    ('Fees and Interest', 'Fees & Interest'),
]


# ---------------------------------------------------------------------------
# Challenge template catalog — used by GET /api/challenges/suggestions
# ---------------------------------------------------------------------------
# Each entry: (product_key, name, challenge_type, bonus_type, bonus_amount,
#              spend_cap, spend_threshold, recurrence, notes)
# recurrence: 'annual' | 'quarterly' | 'once'
# Spend amounts in dollars; bonus_amount in pts/$ or flat pts or # of benefits.
CHALLENGE_TEMPLATES = [
    # ── Hilton Aspire ─────────────────────────────────────────────────────
    ("hilton_aspire", "Free Night Award (annual $30K spend)",
     "annual_threshold", "benefit", 1, None, 30000, "annual",
     "Earn one Free Night Reward Certificate after $30K calendar-year spend. Resets every Jan 1."),
    ("hilton_aspire", "Second Free Night Award ($60K spend)",
     "annual_threshold", "benefit", 1, None, 60000, "annual",
     "Earn a second Free Night Reward Certificate after $60K calendar-year spend. Resets every Jan 1."),

    # ── Hilton Surpass ────────────────────────────────────────────────────
    ("hilton_surpass", "Free Night Certificate (annual $15K spend)",
     "annual_threshold", "benefit", 1, None, 15000, "annual",
     "Earn one Free Night Reward Certificate."),

    # ── Chase Freedom / Freedom Flex rotating 5x ─────────────────────────
    ("chase_freedom", "Q1 Rotating 5x (up to $1,500)",
     "category_rate_cap", "per_dollar", 4, 1500, None, "quarterly",
     "Activate by 3/14. Categories announced each quarter."),
    ("chase_freedom_flex", "Q1 Rotating 5x (up to $1,500)",
     "category_rate_cap", "per_dollar", 4, 1500, None, "quarterly",
     "Activate by 3/14. Same categories as Freedom."),

    # ── Marriott Bonvoy Brilliant ─────────────────────────────────────────
    ("marriott_bonvoy_brilliant", "Earned Choice Award",
     "annual_threshold", "flat", 0, None, 60000, "annual",
     "Choose one: 5 Nightly Upgrade Awards, Free Night up to 85K pts, or 50K bonus points. Resets Jan 1."),

    # ── World of Hyatt ────────────────────────────────────────────────────
    ("world_of_hyatt", "Free Night Certificate (annual $15K spend)",
     "annual_threshold", "benefit", 1, None, 15000, "annual",
     "Earn one Category 1-4 Free Night Certificate."),
    ("world_of_hyatt", "Second Free Night Certificate ($30K total)",
     "annual_threshold", "benefit", 1, None, 30000, "annual",
     "Earn second Cat 1-4 certificate when annual spend hits $30K."),

    # ── Delta SkyMiles Gold ───────────────────────────────────────────────
    ("delta_gold", "$200 Delta Flight Credit ($10K calendar-year spend)",
     "annual_threshold", "flat", 200, None, 10000, "annual",
     "Earn a $200 Delta Flight Credit after $10,000 in calendar-year purchases. Resets every Jan 1."),

    # ── United Quest ──────────────────────────────────────────────────────
    ("united_quest", "Second 10,000-Mile Anniversary Award Discount ($20K spend)",
     "annual_threshold", "flat", 10000, None, 20000, "annual",
     "Earn a second 10,000-mile anniversary award discount after $20K calendar-year spend."),
    ("united_quest", "2 Economy Plus Seat Upgrades ($40K spend)",
     "annual_threshold", "benefit", 2, None, 40000, "annual",
     "Earn 2 Economy Plus seat upgrade certificates after $40K calendar-year spend."),

    # ── Atmos Ascent ──────────────────────────────────────────────────────
    ("atmos_ascent", "Annual Companion Fare",
     "annual_threshold", "flat", 99, None, 6000, "annual",
     "$99 companion fare (+taxes from $23) for saver/main cabin within North America including Hawaii"),

    # ── United Explorer ───────────────────────────────────────────────────
    ("united_explorer", "$100 United TravelBank Credit ($10K spend)",
     "annual_threshold", "flat", 100, None, 10000, "annual",
     "Earn a $100 United TravelBank credit after $10,000 in calendar-year purchases."),
    ("united_explorer", "10K Mile Award Discount ($20K spend)",
     "annual_threshold", "flat", 10000, None, 20000, "annual",
     "Earn a 10,000-mile anniversary award discount after $20,000 in calendar-year purchases."),

    # ── Marriott Bonvoy Boundless ──────────────────────────────────────────
    ("marriott_bonvoy_boundless", "$50 Airline Credit (H1)",
     "annual_threshold", "flat", 50, None, 250, "semi-annual",
     "$50 statement credit after $250 in direct airline purchases Jan 1–Jun 30."),
    ("marriott_bonvoy_boundless", "$50 Airline Credit (H2)",
     "annual_threshold", "flat", 50, None, 250, "semi-annual",
     "$50 statement credit after $250 in direct airline purchases Jul 1–Dec 31."),
    ("marriott_bonvoy_boundless", "Gold Elite Status Upgrade",
     "annual_threshold", "flat", 0, None, 35000, "annual",
     "Marriott Bonvoy Gold Elite status after $35,000 in calendar-year spend."),
]


def _is_sqlite(engine):
    return str(engine.url).startswith('sqlite')


def run_migrations(engine):
    """
    Auto-migrate: safely add any missing columns to existing tables.
    Safe to run on every startup — skips columns that already exist.
    Supports both SQLite and PostgreSQL.
    """
    required_columns = {
        'transactions': [
            ('is_locked',       'BOOLEAN DEFAULT FALSE'),
            ('is_split',        'BOOLEAN DEFAULT FALSE'),
            ('parent_transaction_id', 'INTEGER'),
            ('gcb_tagged',      'BOOLEAN DEFAULT FALSE'),
            ('is_gcb',          'BOOLEAN DEFAULT FALSE'),
            ('points_category', 'VARCHAR(100)'),
            ('card_id',         'INTEGER'),
            ('reviewed_at',       'TIMESTAMP'),
            ('enrichment_source', 'VARCHAR(20)'),
            ('import_hash',       'VARCHAR(64)'),
            ('import_source',     'VARCHAR(20)'),
            ('loan_id',           'INTEGER'),
            ('is_excluded',       'BOOLEAN DEFAULT FALSE'),
            ('content_hash',      'VARCHAR(20)'),
            ('points_earn_override', 'FLOAT'),
            ('spender',           'VARCHAR(100)'),
        ],
        'accounts': [
            ('is_manual', 'BOOLEAN DEFAULT FALSE'),
            ('starting_balance', 'FLOAT DEFAULT 0'),
            ('start_date', 'DATE'),
            ('notes', 'TEXT'),
            ('persistent_account_id', 'VARCHAR(200)'),
            ('institution_id', 'VARCHAR(50)'),
            ('account_hash',   'VARCHAR(16)'),
            ('liability_min_payment',      'FLOAT'),
            ('liability_next_due_date',    'DATE'),
            ('liability_last_statement_bal', 'FLOAT'),
            ('liability_last_payment',     'FLOAT'),
            ('liability_last_payment_date', 'DATE'),
            ('liability_purchase_apr',     'FLOAT'),
            ('product_id',                 'INTEGER'),
        ],
        'plaid_items': [
            ('institution_id',    'VARCHAR(50)'),
            ('last_error_code',   'VARCHAR(100)'),
            ('last_error_message','TEXT'),
            ('last_error_at',     'TIMESTAMP'),
        ],
        'cards': [
            ('account_id', 'INTEGER'),
            ('payment_account_id', 'INTEGER'),
            ('product_id', 'INTEGER'),
            ('ecosystem_id', 'INTEGER'),
        ],
        'loans': [
            ('balance_date',          'DATE'),
            ('remaining_term_months', 'INTEGER'),
            ('property_tax_monthly',  'FLOAT'),
            ('insurance_monthly',     'FLOAT'),
            ('payment_account_id',    'INTEGER'),
            ('payment_due_day',       'INTEGER'),
        ],
        'categorization_rules': [
            ('clean_description', 'VARCHAR(500)'),
            ('priority_order', 'INTEGER DEFAULT 0'),
        ],
        'transaction_splits': [
            ('action', 'VARCHAR(50)'),
        ],
        'points_ecosystems': [
            ('eco_type',           'VARCHAR(20)'),
            ('conservative_basis', 'VARCHAR(200)'),
        ],
        'points_categories': [
            ('parent_key', 'VARCHAR(100)'),
        ],
        'spend_challenges': [
            ('activation_date',  'DATE'),
            ('name',             "VARCHAR(200) DEFAULT ''"),
            ('challenge_type',   "VARCHAR(30) DEFAULT 'sub'"),
            ('bonus_type',       "VARCHAR(20) DEFAULT 'flat'"),
            ('bonus_amount',     'FLOAT DEFAULT 0'),
            ('spend_cap',        'FLOAT'),
            ('spend_threshold',  'FLOAT'),
            ('spender_filter',   'VARCHAR(100)'),
            ('current_spend',    'FLOAT DEFAULT 0'),
            ('bonus_unlocked',   'BOOLEAN DEFAULT FALSE'),
            ('is_active',        'BOOLEAN DEFAULT TRUE'),
            ('notes',            'TEXT'),
            ('created_at',       'TIMESTAMP DEFAULT NOW()'),
        ],
        'card_product_rewards': [
            ('reward_type', "VARCHAR(50) DEFAULT 'fixed'"),
        ],
        'card_benefits': [
            ('trigger_category', 'VARCHAR(100)'),
            ('notes',            'TEXT'),
            ('tracking_type',    'VARCHAR(20)'),  # no default — NULL means "needs classifying", see backfill below
        ],
        'benefit_usage': [
            ('card_id',    'INTEGER'),
            ('confirmed',  'BOOLEAN DEFAULT FALSE'),
            ('notes',      'TEXT'),
            ('updated_at', 'TIMESTAMP DEFAULT NOW()'),
        ],
    }

    if _is_sqlite(engine):
        _run_migrations_sqlite(engine, required_columns)
    else:
        _run_migrations_pg(engine, required_columns)


def _run_migrations_sqlite(engine, required_columns):
    import sqlite3
    db_path = str(engine.url).replace('sqlite:///', '')
    conn = sqlite3.connect(db_path)
    try:
        for table, columns in required_columns.items():
            cursor = conn.execute(f'PRAGMA table_info({table})')
            existing = {row[1] for row in cursor.fetchall()}
            for col_name, col_def in columns:
                if col_name not in existing:
                    conn.execute(f'ALTER TABLE {table} ADD COLUMN {col_name} {col_def}')
                    print(f'  Migration: added {table}.{col_name}')
        try:
            cursor = conn.execute('PRAGMA table_info(transactions)')
            cols = {row[1] for row in cursor.fetchall()}
            if 'is_gcb' in cols and 'gcb_tagged' in cols:
                conn.execute('UPDATE transactions SET is_gcb = gcb_tagged WHERE is_gcb = 0 AND gcb_tagged = 1')
                print('  Migration: copied gcb_tagged → is_gcb')
        except Exception:
            pass
        try:
            cursor = conn.execute('PRAGMA table_info(budget_targets)')
            bt_cols = {row[1] for row in cursor.fetchall()}
            if bt_cols and 'year' not in bt_cols:
                conn.execute('DROP TABLE IF EXISTS budget_targets')
                print('  Migration: dropped old budget_targets table (will be recreated)')
        except Exception:
            pass
        # One-time classification of card_benefits.tracking_type ('periodic' vs
        # 'by_use') for rows that predate the column — see _run_migrations_pg for
        # the rationale. IS NULL guard makes it safe to run every startup.
        try:
            cursor = conn.execute('PRAGMA table_info(card_benefits)')
            cb_cols = {row[1] for row in cursor.fetchall()}
            if 'tracking_type' in cb_cols:
                conn.execute("""
                    UPDATE card_benefits SET tracking_type = 'by_use'
                    WHERE tracking_type IS NULL
                      AND (
                        reset_frequency LIKE 'every_4%'
                        OR reset_frequency LIKE 'every_5%'
                        OR benefit_name LIKE '%per qualifying%'
                        OR benefit_name LIKE '%per claim%'
                        OR benefit_name LIKE '%per use%'
                        OR benefit_name LIKE '%per stay%'
                        OR benefit_name LIKE '%per night%'
                        OR benefit_name LIKE '%every 4%'
                        OR benefit_name LIKE '%every 5%'
                      )
                """)
                conn.execute(
                    "UPDATE card_benefits SET tracking_type = 'periodic' WHERE tracking_type IS NULL"
                )
                print('  Migration: classified card_benefits.tracking_type')
        except Exception:
            pass
        # Normalize account_type to Title Case
        try:
            type_map = [
                ('checking',    'Checking'),
                ('savings',     'Savings'),
                ('brokerage',   'Brokerage'),
                ('investment',  'Investment'),
                ('credit card', 'Credit Card'),
                ('credit',      'Credit Card'),
                ('loan',        'Loan'),
                ('other',       'Other'),
            ]
            for old_val, new_val in type_map:
                conn.execute(
                    "UPDATE accounts SET account_type = ? WHERE LOWER(account_type) = ?",
                    (new_val, old_val)
                )
            print('  Migration: normalized account_type casing')
        except Exception:
            pass
        # ── Category remap (consolidation) ──────────────────────────────────
        try:
            for old, new in _CAT_REMAP:
                conn.execute(
                    "UPDATE transactions SET category_auto = ? WHERE category_auto = ?", (new, old))
                conn.execute(
                    "UPDATE transactions SET category_manual = ? WHERE category_manual = ?", (new, old))
                conn.execute(
                    "UPDATE budget_targets SET category = ? WHERE category = ?", (new, old))
                conn.execute(
                    "UPDATE categorization_rules SET set_category = ? WHERE set_category = ?", (new, old))
            print('  Migration: remapped consolidated categories')
        except Exception:
            pass
        # ── Reclassify CC credits: positive amounts on CC accounts are Expense, not Income ──
        try:
            conn.execute("""
                UPDATE transactions
                SET action = 'Expense'
                WHERE action = 'Income'
                  AND amount > 0
                  AND (is_locked = 0 OR is_locked IS NULL)
                  AND account_id IN (
                      SELECT id FROM accounts
                      WHERE LOWER(account_type) IN ('credit', 'credit card')
                  )
                  AND UPPER(description_raw) NOT LIKE '%PAYROLL%'
                  AND UPPER(description_raw) NOT LIKE '%DIR DEP%'
                  AND UPPER(description_raw) NOT LIKE '%DIRECT DEP%'
            """)
            print('  Migration: reclassified CC credits from Income → Expense')
        except Exception:
            pass
        try:
            conn.execute('DROP TABLE IF EXISTS merchant_overrides')
        except Exception:
            pass
        conn.commit()
    finally:
        conn.close()


def _run_migrations_pg(engine, required_columns):
    from sqlalchemy import text, inspect
    insp = inspect(engine)
    with engine.begin() as conn:
        for table, columns in required_columns.items():
            if not insp.has_table(table):
                continue
            existing = {c['name'] for c in insp.get_columns(table)}
            for col_name, col_def in columns:
                if col_name not in existing:
                    conn.execute(text(f'ALTER TABLE {table} ADD COLUMN {col_name} {col_def}'))
                    print(f'  Migration: added {table}.{col_name}')
        if insp.has_table('transactions'):
            existing = {c['name'] for c in insp.get_columns('transactions')}
            if 'is_gcb' in existing and 'gcb_tagged' in existing:
                conn.execute(text(
                    'UPDATE transactions SET is_gcb = gcb_tagged WHERE is_gcb = FALSE AND gcb_tagged = TRUE'
                ))
                print('  Migration: copied gcb_tagged → is_gcb')
        # Backfill institution_id onto accounts that still have NULL.
        # Copies from the linked plaid_items row. Safe to run every startup
        # (IS NULL guard makes it idempotent). Preserves institution_id on
        # severed accounts (plaid_item_id=NULL) — those were already written.
        if insp.has_table('accounts') and insp.has_table('plaid_items'):
            existing_acct_cols = {c['name'] for c in insp.get_columns('accounts')}
            if 'institution_id' in existing_acct_cols:
                result = conn.execute(text(
                    "UPDATE accounts a"
                    " SET institution_id = pi.institution_id"
                    " FROM plaid_items pi"
                    " WHERE a.plaid_item_id = pi.item_id"
                    "   AND a.institution_id IS NULL"
                    "   AND pi.institution_id IS NOT NULL"
                ))
                if result.rowcount:
                    print(f'  Migration: backfilled institution_id on {result.rowcount} account(s)')
        # One-time classification of card_benefits.tracking_type ('periodic' vs
        # 'by_use') for rows that predate the column. IS NULL guard makes this
        # safe to run every startup without clobbering later explicit choices
        # (heuristic-assigned or user-edited via the UI).
        # NOTE: uses `conn` (not `insp`, which reflects via its own connection
        # and won't see this same transaction's uncommitted ALTER TABLE) —
        # just attempt it and swallow the error if the column truly isn't there.
        if insp.has_table('card_benefits'):
            try:
                result = conn.execute(text("""
                    UPDATE card_benefits SET tracking_type = 'by_use'
                    WHERE tracking_type IS NULL
                      AND (
                        reset_frequency ILIKE 'every_4%'
                        OR reset_frequency ILIKE 'every_5%'
                        OR benefit_name ILIKE '%per qualifying%'
                        OR benefit_name ILIKE '%per claim%'
                        OR benefit_name ILIKE '%per use%'
                        OR benefit_name ILIKE '%per stay%'
                        OR benefit_name ILIKE '%per night%'
                        OR benefit_name ILIKE '%every 4%'
                        OR benefit_name ILIKE '%every 5%'
                      )
                """))
                if result.rowcount:
                    print(f'  Migration: classified {result.rowcount} card_benefit(s) as by_use')
                result = conn.execute(text(
                    "UPDATE card_benefits SET tracking_type = 'periodic' WHERE tracking_type IS NULL"
                ))
                if result.rowcount:
                    print(f'  Migration: classified {result.rowcount} card_benefit(s) as periodic')
            except Exception:
                pass
        # Normalize account_type to Title Case
        if insp.has_table('accounts'):
            type_map = [
                ('checking',    'Checking'),
                ('savings',     'Savings'),
                ('brokerage',   'Brokerage'),
                ('investment',  'Investment'),
                ('credit card', 'Credit Card'),
                ('credit',      'Credit Card'),
                ('loan',        'Loan'),
                ('other',       'Other'),
            ]
            for old_val, new_val in type_map:
                conn.execute(text(
                    "UPDATE accounts SET account_type = :new WHERE LOWER(account_type) = :old"
                ), {'new': new_val, 'old': old_val})
            print('  Migration: normalized account_type casing')
        # ── Category remap (consolidation) ──────────────────────────────────
        if insp.has_table('transactions'):
            for old, new in _CAT_REMAP:
                conn.execute(text(
                    "UPDATE transactions SET category_auto = :new WHERE category_auto = :old"
                ), {'new': new, 'old': old})
                conn.execute(text(
                    "UPDATE transactions SET category_manual = :new WHERE category_manual = :old"
                ), {'new': new, 'old': old})
            print('  Migration: remapped transaction categories')
        if insp.has_table('budget_targets'):
            for old, new in _CAT_REMAP:
                conn.execute(text(
                    "UPDATE budget_targets SET category = :new WHERE category = :old"
                ), {'new': new, 'old': old})
            print('  Migration: remapped budget_targets categories')
        if insp.has_table('categorization_rules'):
            for old, new in _CAT_REMAP:
                conn.execute(text(
                    "UPDATE categorization_rules SET set_category = :new WHERE set_category = :old"
                ), {'new': new, 'old': old})
            print('  Migration: remapped categorization_rules set_category')
        # ── Reclassify CC credits: positive amounts on CC accounts are Expense, not Income ──
        # Drop stale NOT NULL constraints on spend_challenges columns that exist
        # only in older Railway DB schema versions but are absent from the current
        # model.  Any such column with NOT NULL will block every INSERT.
        if insp.has_table('spend_challenges'):
            _sc_cols = {c['name']: c for c in insp.get_columns('spend_challenges')}
            # Columns that may exist in old schema with NOT NULL but are no longer
            # part of the model — make them nullable so INSERTs succeed.
            # Known model columns — anything ELSE in the DB is a legacy relic.
            # Drop NOT NULL on every such relic so INSERTs from the current model
            # always succeed regardless of which old schema the DB was created from.
            _sc_model_cols = {
                'id', 'card_id', 'name', 'challenge_type',
                'start_date', 'end_date', 'activation_date',
                'bonus_type', 'bonus_amount',
                'spend_cap', 'spend_threshold',
                'current_spend', 'bonus_unlocked',
                'is_active', 'notes', 'created_at',
            }
            for _col_name, _col_info in _sc_cols.items():
                if _col_name not in _sc_model_cols and not _col_info.get('nullable', True):
                    conn.execute(text(
                        f'ALTER TABLE spend_challenges ALTER COLUMN {_col_name} DROP NOT NULL'
                    ))
                    print(f'  Migration: dropped NOT NULL on spend_challenges.{_col_name}')
        if insp.has_table('transactions') and insp.has_table('accounts'):
            conn.execute(text("""
                UPDATE transactions
                SET action = 'Expense'
                WHERE action = 'Income'
                  AND amount > 0
                  AND (is_locked = FALSE OR is_locked IS NULL)
                  AND account_id IN (
                      SELECT id FROM accounts
                      WHERE LOWER(account_type) IN ('credit', 'credit card')
                  )
                  AND UPPER(description_raw) NOT LIKE '%PAYROLL%'
                  AND UPPER(description_raw) NOT LIKE '%DIR DEP%'
                  AND UPPER(description_raw) NOT LIKE '%DIRECT DEP%'
            """))
            print('  Migration: reclassified CC credits from Income → Expense')
        # merchant_overrides was a dead system — never populated by any UI path.
        # Superseded entirely by CategorizationRule (rules are the only source
        # of automatic classification now); safe to drop unconditionally.
        if insp.has_table('merchant_overrides'):
            conn.execute(text('DROP TABLE IF EXISTS merchant_overrides'))
            print('  Migration: dropped unused merchant_overrides table')

# Database initialization
def init_db(database_url=None):
    """Initialize the database with schema. Reads DATABASE_URL from env, falls back to SQLite."""
    if database_url is None:
        database_url = os.getenv('DATABASE_URL', 'sqlite:///./finance.db')
    # Railway/Heroku may provide postgres:// but SQLAlchemy 2.x needs postgresql://
    if database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql://', 1)
    engine = create_engine(database_url, echo=False)
    Base.metadata.create_all(engine)
    run_migrations(engine)
    # Re-run create_all to pick up any tables dropped by migration (e.g. budget_targets schema change)
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine)
    return engine, SessionLocal


def seed_categories(session):
    """
    Seed / refresh the category list.
    Adds new categories if missing; marks removed categories inactive.
    Safe to re-run on every startup.
    """
    categories_data = [
        # ── Personal expense categories ───────────────────────────────────────
        ("Groceries",              None, "expense",  1),
        ("Dining",                 None, "expense",  2),
        ("Transportation",         None, "expense",  3),  # gas, parking, transit, rides
        ("Housing",                None, "expense",  4),  # rent / mortgage cash payment
        ("Utilities",              None, "expense",  5),  # phone, internet, water, electricity
        ("Healthcare",             None, "expense",  6),
        ("Insurance",              None, "expense",  7),
        ("Vehicle",                None, "expense",  8),  # maintenance, registration
        ("Fitness",                None, "expense",  9),
        ("Self Care",              None, "expense", 10),  # grooming, dry cleaning, spa
        ("Clothing",               None, "expense", 11),
        ("Electronics",            None, "expense", 12),
        ("Streaming",              None, "expense", 13),  # subscriptions
        ("Travel",                 None, "expense", 14),
        ("Home",                   None, "expense", 15),  # furniture, repairs, maintenance
        ("Kids",                   None, "expense", 16),
        ("Entertainment",          None, "expense", 17),  # events, leisure, books, lottery
        ("Gifts",                  None, "expense", 18),  # presents
        ("For Others",             None, "expense", 19),  # financial support for family/others
        ("Education",              None, "expense", 20),  # incl. lessons, tutoring, studies
        ("Fees & Interest",        None, "expense", 21),
        ("Other",                  None, "expense", 22),
        # ── Both income and expense (use Action to distinguish) ───────────────
        ("Business",               None, "both",    23),  # GCB + future ventures; Action = Income or Expense
        ("Investment Gain (Loss)", None, "both",    24),  # gains, losses, interest income
        # ── Income ───────────────────────────────────────────────────────────
        ("Work",                   None, "income",   1),  # salary / payroll
        # ── Special / system ─────────────────────────────────────────────────
        ("Unclassified",           None, "both",   100),
        ("Transfer",               None, "both",   101),
    ]

    active_names = {row[0] for row in categories_data}

    for name, parent, cat_type, order in categories_data:
        existing = session.query(Category).filter_by(name=name).first()
        if not existing:
            session.add(Category(
                name=name,
                parent_category=parent,
                category_type=cat_type,
                display_order=order,
                is_active=True,
            ))
        else:
            # Re-activate in case it was marked inactive by a previous migration
            existing.is_active = True
            existing.display_order = order
            existing.category_type = cat_type

    # Deactivate any categories that are no longer in the canonical list
    for cat in session.query(Category).all():
        if cat.name not in active_names:
            cat.is_active = False

    session.commit()


if __name__ == "__main__":
    # Test database creation
    engine, SessionLocal = init_db()
    session = SessionLocal()
    seed_categories(session)
    session.close()
    print("Database initialized successfully!")

def seed_points_categories(session):
    """Seed points earning categories with a two-level hierarchy.

    Format: (name, display_order, parent_key)

    L1 categories have parent_key=None.
    L2 (brand-specific) categories carry their L1 parent name so the points
    calculation engine can fall back: L2 rate → L1 rate → base rate.
    """
    cats = [
        # ── L1: broad spend categories ────────────────────────────────────
        ("Dining",                 1,  None),
        ("Food Delivery",          2,  None),
        ("Groceries",              3,  None),
        ("Airlines",               4,  None),
        ("Ground Transportation",  5,  None),
        ("Hotels",                 6,  None),
        ("Gas Stations",           7,  None),
        ("Car Rental",             8,  None),
        ("Wholesale Clubs",        9,  None),
        ("Online Shopping",        16, None),
        ("Drugstore",              13, None),
        ("Streaming",              26, None),
        ("Chase Travel",           25, None),   # booking-channel L1 (Chase portal)
        ("Spa & Salon",            24, None),
        ("Transit",                36, None),
        ("Home Improvement",       37, None),
        ("Fitness & Gyms",         38, None),
        ("Live Entertainment",     39, None),
        # ── L2: hotel brands → Hotels ─────────────────────────────────────
        ("Hilton",                 17, "Hotels"),
        ("Marriott",               18, "Hotels"),
        ("Hyatt",                  19, "Hotels"),
        ("IHG",                    33, "Hotels"),
        # ── L2: airline brands → Airlines ─────────────────────────────────
        ("United",                 27, "Airlines"),
        ("Delta",                  28, "Airlines"),
        ("American Airlines",      29, "Airlines"),
        ("Southwest",              30, "Airlines"),
        ("JetBlue",                31, "Airlines"),
        ("Alaska Airlines",        32, "Airlines"),
        ("Hawaiian Airlines",      33, "Airlines"),
        # ── L2: rideshare brands → Ground Transportation ──────────────────
        ("Rideshare: Lyft",        14, "Ground Transportation"),
        ("Rideshare: Uber",        15, "Ground Transportation"),
        # ── L2: grocery/retail brands ─────────────────────────────────────
        ("Walmart",                10, "Groceries"),   # Visa/MC classify as grocery
        ("Target",                 11, "Groceries"),   # Same MCC as grocery at most issuers
        ("Amazon",                 12, "Online Shopping"),
        # ── L2: United co-branded categories ──────────────────────────────
        ("United Purchases",       34, "Airlines"),   # United-operated flights, bags, upgrades, etc.
        # ── L2: luxury hotel channel → Hotels ─────────────────────────────
        ("Renowned Hotels & Resorts", 35, "Hotels"),  # Prepaid bookings via United's Renowned Hotels
        # ── L2: co-branded retail (no meaningful L1 parent) ───────────────
        ("Best Buy",               20, None),
        ("Marshalls",              21, None),
        ("West Elm",               22, None),
        ("C&B",                    23, None),
        # ── L2: EV / utilities → Ground Transportation ────────────────────
        ("EV Charging",            40, "Ground Transportation"),
        # ── L2: cable → Streaming ─────────────────────────────────────────
        ("Cable",                  41, "Streaming"),
        # ── L1: person-to-person payments (never earn points, see _NON_EARNING_CATS) ──
        ("P2P Payments",           42, None),
    ]
    for name, order, parent in cats:
        existing = session.query(PointsCategory).filter_by(name=name).first()
        if not existing:
            session.add(PointsCategory(name=name, display_order=order, parent_key=parent))
        else:
            # Always refresh parent_key so hierarchy corrections take effect on deploy
            existing.parent_key = parent
            existing.display_order = order
    session.commit()


def seed_points_ecosystems(session):
    """Seed all points/miles ecosystems with valuations."""
    ecosystems = [
        ("Chase UR", "Ultimate Rewards", "Flexible", 1.0, False, "Points Boost variable rate (1.25cpp portal guarantee removed Oct 2025)"),
        ("Amex MR", "Membership Rewards", "Flexible", 1.0, False, "Transfer partner average"),
        ("Citi ThankYou", "ThankYou Points", "Flexible", 1.0, False, "Transfer partner average"),
        ("Capital One Miles", "Capital One Miles", "Flexible", 1.0, False, "Transfer/portal"),
        ("Hilton Honors", "Hilton Honors Points", "Hotel", 0.5, False, "Standard night avg"),
        ("Marriott Bonvoy", "Marriott Bonvoy Points", "Hotel", 0.7, False, "Standard night avg"),
        ("World of Hyatt", "World of Hyatt Points", "Hotel", 1.7, False, "Standard night avg"),
        ("IHG Rewards", "IHG One Rewards Points", "Hotel", 0.5, False, "Standard night avg"),
        ("Delta SkyMiles", "Delta SkyMiles", "Airline", 1.2, False, "Domestic economy avg"),
        ("United MileagePlus", "United MileagePlus Miles", "Airline", 1.0, False, "Domestic economy avg"),
        ("AA AAdvantage", "AAdvantage Miles", "Airline", 1.0, False, "Domestic economy avg"),
        ("Southwest RR", "Rapid Rewards Points", "Airline", 1.0, False, "Wanna Get Away avg"),
        ("JetBlue TrueBlue", "TrueBlue Points", "Airline", 1.0, False, "Domestic economy avg"),
        ("Alaska Mileage Plan", "Alaska Miles", "Airline", 1.0, False, "Domestic economy avg"),
        ("Cash Back", "Cash Back", "Cash", 1.0, True, "1:1 cash value"),
        ("Discover Cashback", "Cashback Bonus", "Cash", 1.0, True, "1:1 cash value"),
        ("Best Buy Rewards", "Best Buy Reward Certificates", "Cash", 1.0, True, "1:1 cash value"),
        ("Fidelity Rewards", "Fidelity Rewards", "Cash", 1.0, True, "1:1 cash into brokerage"),
        ("Amazon Rewards", "Amazon Points", "Cash", 1.0, True, "1:1 at Amazon"),
        ("Walmart Rewards", "Walmart Rewards", "Cash", 1.0, True, "1:1 at Walmart"),
        ("Target Circle", "Target Circle Earnings", "Cash", 1.0, True, "1:1 at Target"),
        ("Costco Rewards", "Costco Cash Back", "Cash", 1.0, True, "Annual check"),
        ("Apple Cash", "Apple Cash", "Cash", 1.0, True, "1:1 cash value"),
        ("Atmos Rewards", "Atmos Rewards Points", "Airline", 1.5, False, "TPG valuation"),
    ]
    for name, currency, eco_type, cons_cpp, is_cash, basis in ecosystems:
        existing = session.query(PointsEcosystem).filter_by(name=name).first()
        if not existing:
            session.add(PointsEcosystem(
                name=name, currency_name=currency, eco_type=eco_type,
                conservative_cpp=cons_cpp, your_cpp=cons_cpp,
                is_cash_back=is_cash, conservative_basis=basis,
            ))
        else:
            # Always refresh descriptive/factual fields.
            # conservative_cpp is updated from seed (floor value from card issuer data).
            # your_cpp is intentionally NOT touched — that's the user's personal valuation,
            # EXCEPT when it still equals the old default (meaning user never customised it),
            # in which case we bump it to the new seed value.
            existing.currency_name = currency
            existing.eco_type = eco_type
            existing.is_cash_back = is_cash
            old_cons = float(existing.conservative_cpp or 0)
            existing.conservative_cpp = cons_cpp
            existing.conservative_basis = basis
            # Bump your_cpp if user hasn't diverged from the previous conservative default.
            if abs(float(existing.your_cpp or 0) - old_cons) < 0.001:
                existing.your_cpp = cons_cpp
    session.commit()


_BY_USE_NAME_MARKERS = (
    'per qualifying', 'per claim', 'per use', 'per stay', 'per night',
    'every 4', 'every 5',
)


def _infer_tracking_type(benefit_name: str, reset_frequency: str) -> str:
    """Classify a benefit as 'periodic' (use-it-or-lose-it each cycle — gets the
    checkbox/counter usage tracker) or 'by_use' (Global Entry every-4-yrs credit,
    per-stay/per-claim credits — doesn't expire on a cadence, no tracker/alerts).

    Some benefits are stored with a misleading reset_frequency (e.g. a "every 4 yrs"
    Global Entry credit tagged 'annual' in the seed data below) — the name-based
    markers catch those cases the frequency field alone would miss.
    """
    if (reset_frequency or '').lower() in ('every_4.5_years', 'every_4_years', 'every_5_years'):
        return 'by_use'
    name_lower = (benefit_name or '').lower()
    if any(marker in name_lower for marker in _BY_USE_NAME_MARKERS):
        return 'by_use'
    return 'periodic'


def seed_card_products(session):
    """
    Seed the full card product catalog with earning rates and benefits.
    Runs on every startup — skips products that already exist (by product_key).
    Does NOT overwrite existing data (so user edits via UI are preserved).
    """
    # Build category lookup
    cat_map = {c.name: c.id for c in session.query(PointsCategory).all()}
    eco_map = {e.name: e.id for e in session.query(PointsEcosystem).all()}

    # Product definitions: (product_key, card_name, ecosystem_name, status, annual_fee, benefits, earning_rates)
    # earning_rates: dict of {category_name: additional_points_above_base} + {'_base': base_rate}
    # benefits: list of (name, amount, frequency, trigger_category)
    products = [
        ("chase_sapphire_preferred", "Chase Sapphire Preferred", "Chase UR", "active", 95, [
            ("$50 Chase Travel Hotel Credit", 50, "annual", "Chase Travel"),
            ("DashPass Membership", 0, "annual", "Food Delivery"),
            ("$10/mo DoorDash Credit", 120, "annual", "Food Delivery"),
        ], {"_base": 1, "Chase Travel": 4, "Dining": 2, "Food Delivery": 2, "Groceries": 2,
            "Streaming": 2, "Airlines": 1, "Ground Transportation": 1, "Hotels": 1,
            "Car Rental": 1, "Rideshare: Lyft": 4}),

        ("chase_sapphire_reserve", "Chase Sapphire Reserve", "Chase UR", "not_held", 795, [
            ("$300 Travel Credit", 300, "annual", None),
            ("$500 The Edit Hotel Credit (2x $250)", 500, "calendar_year", "Chase Travel"),
            ("$250 Chase Travel Hotel Credit", 250, "calendar_year", "Chase Travel"),
            ("$300 Exclusive Tables Dining Credit (2x $150 semi-annual)", 300, "calendar_year", "Dining"),
            ("$300 StubHub/Viagogo Credit (2x $150 semi-annual)", 300, "calendar_year", None),
            ("$25/mo DoorDash Credit", 300, "annual", "Food Delivery"),
            ("DashPass Membership", 0, "annual", "Food Delivery"),
            ("$10/mo Lyft Credit", 120, "annual", "Rideshare: Lyft"),
            ("$10/mo Peloton Credit", 120, "annual", None),
            ("Apple TV+ & Apple Music (through June 2027)", 288, "annual", None),
            ("Global Entry/TSA PreCheck/NEXUS Credit", 120, "every_4.5_years", None),
            ("Priority Pass Lounge Access", 0, "annual", None),
            ("IHG Platinum Elite Status (through 2027)", 0, "annual", None),
        ], {"_base": 1, "Chase Travel": 7, "Dining": 2, "Airlines": 3, "Hotels": 3,
            "Rideshare: Lyft": 4}),

        ("chase_freedom", "Chase Freedom", "Chase UR", "active", 0, [
            ("5x Rotating Categories (activate quarterly, cap $1,500/qtr)", 0, "quarterly", None),
        ],
         {"_base": 1}),

        ("chase_freedom_unlimited", "Chase Freedom Unlimited", "Chase UR", "active", 0, [],
         {"_base": 1.5, "Dining": 1.5, "Drugstore": 1.5, "Chase Travel": 3.5, "Rideshare: Lyft": 3.5}),

        ("chase_freedom_flex", "Chase Freedom Flex", "Chase UR", "active", 0, [],
         {"_base": 1, "Dining": 2, "Food Delivery": 2, "Drugstore": 2}),

        ("chase_ink_preferred", "Chase Ink Business Preferred", "Chase UR", "not_held", 95, [],
         {"_base": 1}),

        ("chase_ink_unlimited", "Chase Ink Business Unlimited", "Chase UR", "not_held", 0, [],
         {"_base": 1.5}),

        ("chase_amazon_prime", "Amazon Prime Rewards Visa", "Cash Back", "active", 0, [],
         {"_base": 1, "Amazon": 4, "Groceries": 1, "Dining": 1, "Gas Stations": 1}),

        ("amex_platinum", "Amex Platinum", "Amex MR", "active", 695, [
            ("$200 Airline Fee Credit", 200, "calendar_year", "Airlines"),
            ("$200 Hotel Credit (FHR/THC)", 200, "calendar_year", "Hotels"),
            ("$200 Uber Cash", 200, "annual", "Rideshare: Uber"),
            ("$155 Walmart+ Credit", 155, "annual", "Walmart"),
            ("$240 Digital Entertainment", 240, "annual", None),
            ("$189 CLEAR Plus Credit", 189, "annual", None),
            ("$100 Saks Credit", 100, "annual", None),
            ("Global Entry/TSA PreCheck", 100, "every_4.5_years", None),
            ("Centurion Lounge Access", 0, "annual", None),
        ], {"_base": 1, "Airlines": 4}),

        ("amex_gold", "Amex Gold", "Amex MR", "not_held", 325, [
            ("$120 Uber Cash", 10, "monthly", "Rideshare: Uber"),
            ("$120 Dining Credit", 10, "monthly", "Dining"),
            ("$100 Resy Credit", 50, "semi-annual", "Dining"),
            ("$84 Dunkin' Credit", 7, "monthly", "Dining"),
        ], {"_base": 1, "Dining": 3, "Food Delivery": 3, "Groceries": 3, "Airlines": 2}),

        ("amex_green", "Amex Green", "Amex MR", "not_held", 150, [
            ("$189 CLEAR Plus Credit", 189, "annual", None),
            ("$100 LoungeBuddy Credit", 100, "annual", None),
        ], {"_base": 1, "Dining": 2, "Airlines": 2, "Ground Transportation": 2}),

        ("amex_blue_cash_preferred", "Blue Cash Preferred", "Cash Back", "not_held", 95, [],
         {"_base": 1, "Groceries": 5, "Streaming": 5, "Gas Stations": 2, "Transit": 2}),

        ("hilton_aspire", "Hilton Honors Aspire", "Hilton Honors", "active", 550, [
            ("$400 Hilton Resort Credit ($200 semi-annual, Hilton resort properties)", 200, "semi-annual", "Hilton"),
            ("$200 Flight Credit ($50/quarter, direct or Amex Travel)", 50, "quarterly", "Airlines"),
            ("$209 CLEAR+ Credit (annual)", 209, "annual", None),
            ("Free Night Award (annual, at card anniversary)", 0, "annual", "Hilton"),
            ("$100 Waldorf/Conrad On-Property Credit (per qualifying 2-night+ stay)", 100, "annual", "Hilton"),
            ("Complimentary Hilton Diamond Status", 0, "annual", None),
        ], {"_base": 3, "Dining": 4, "Airlines": 4, "Car Rental": 4, "Hilton": 11}),

        ("hilton_surpass", "Hilton Honors Surpass", "Hilton Honors", "not_held", 150, [
            ("Gold Status", 0, "annual", None),
            ("Free Weekend Night Cert (after $15k)", 0, "annual", "Hilton"),
        ], {"_base": 3, "Dining": 3, "Groceries": 3, "Gas Stations": 3, "Hilton": 9}),

        ("hilton_honors", "Hilton Honors Card", "Hilton Honors", "active", 0, [
            ("Silver Status", 0, "annual", None),
        ], {"_base": 3, "Dining": 2, "Groceries": 2, "Gas Stations": 2, "Hilton": 4}),

        ("hyatt_personal", "World of Hyatt Card", "World of Hyatt", "active", 95, [
            ("Discoverist Status", 0, "annual", None),
            ("Free Night Cert (up to Cat 4)", 0, "annual", "Hyatt"),
            ("Free Night Cert (after $15k)", 0, "annual", "Hyatt"),
        ], {"_base": 1, "Dining": 1, "Spa & Salon": 1, "Hyatt": 3}),

        ("marriott_bonvoy_brilliant", "Marriott Bonvoy Brilliant® American Express® Card", "Marriott Bonvoy", "active", 650, [
            ("$25/mo Dining Credit ($300/yr; statement credit at restaurants worldwide)", 25, "monthly", "Dining"),
            ("Free Night Award (annual, up to 85K Bonvoy pts; issued on renewal)", 0, "annual", "Hotels"),
            ("Marriott Bonvoy Platinum Elite Status", 0, "annual", None),
            ("25 Elite Night Credits", 0, "annual", None),
            ("Priority Pass Select (unlimited visits + 2 guests)", 0, "annual", None),
            ("Global Entry/TSA PreCheck Credit (every 4 yrs; up to $120)", 120, "annual", None),
            ("$100 On-Property Credit (St. Regis/Ritz-Carlton; per qualifying 2-night+ stay)", 100, "annual", "Hotels"),
            ("Cell Phone Protection (up to $800/claim, $50 deductible)", 0, "annual", None),
        ], {"_base": 2, "Dining": 1, "Airlines": 1, "Hotels": 4}),

        ("marriott_bonvoy_boundless", "Marriott Bonvoy Boundless® Credit Card", "Marriott Bonvoy", "active", 95, [
            ("Anniversary Free Night Award (up to 35K pts; can top off with 25K more)", 0, "annual", "Marriott"),
            ("Marriott Bonvoy Silver Elite Status", 0, "annual", None),
            ("15 Elite Night Credits", 0, "annual", None),
            ("DashPass Membership (1 yr complimentary; activate by 12/31/2027)", 0, "annual", "Food Delivery"),
        # Earning: 6x Marriott (2 base + 4 bonus); 3x Groceries/Gas Stations/Dining (2 base + 1 bonus,
        # combined $6,000/yr cap across the three categories); 2x base on everything else.
        ], {"_base": 2, "Marriott": 4, "Groceries": 1, "Gas Stations": 1, "Dining": 1}),

        ("citi_custom_cash", "Citi Custom Cash® Card", "Citi ThankYou", "active", 0, [
            ("Citi Travel Portal Bonus (5x Hotels/Cars/Attractions)", 0, "annual", None),
         ], {"_base": 1, "_auto_top": [
            # Earn 5x (4 additional above 1x base) on whichever eligible category has
            # highest spend each billing cycle, capped at $500/cycle.
            "Dining", "Gas Stations", "Groceries", "Airlines", "Transit",
            "Streaming", "Drugstore", "Home Improvement", "Fitness & Gyms", "Live Entertainment",
         ]}),

        ("citi_double_cash", "Citi Double Cash", "Citi ThankYou", "active", 0, [],
         {"_base": 2}),

        ("citi_premier", "Citi Premier", "Citi ThankYou", "not_held", 95, [],
         {"_base": 1, "Dining": 2, "Groceries": 2, "Airlines": 2, "Gas Stations": 2, "Hotels": 2}),

        ("best_buy_card", "Best Buy Credit Card", "Best Buy Rewards", "active", 0, [],
         {"_base": 1, "Dining": 1, "Food Delivery": 1, "Groceries": 1, "Gas Stations": 2, "Best Buy": 4}),

        ("fidelity_rewards", "Fidelity Rewards Visa", "Fidelity Rewards", "active", 0, [],
         {"_base": 2}),

        ("discover_it", "Discover it Cash Back", "Discover Cashback", "active", 0, [],
         {"_base": 1}),

        ("apple_card", "Apple Card", "Apple Cash", "not_held", 0, [],
         {"_base": 1, "Apple": 2}),

        ("costco_anywhere", "Costco Anywhere Visa", "Costco Rewards", "not_held", 0, [],
         {"_base": 1, "Dining": 2, "Gas Stations": 3, "Wholesale Clubs": 1}),

        ("capital_one_venture_x", "Capital One Venture X", "Capital One Miles", "not_held", 395, [
            ("$300 Travel Credit", 300, "annual", None),
            ("10,000 Anniversary Miles", 0, "annual", None),
            ("Priority Pass + Plaza Premium", 0, "annual", None),
        ], {"_base": 2, "Airlines": 3, "Hotels": 8}),

        ("capital_one_venture", "Capital One Venture", "Capital One Miles", "not_held", 95, [],
         {"_base": 2, "Hotels": 3}),

        ("capital_one_savor_one", "Capital One SavorOne", "Cash Back", "not_held", 0, [],
         {"_base": 1, "Dining": 2, "Groceries": 2, "Streaming": 2, "Live Entertainment": 2}),

        ("delta_gold", "Delta SkyMiles® Gold American Express Card", "Delta SkyMiles", "active", 150, [
            # Auto-trigger benefits — no spend required
            ("Free First Checked Bag (cardmember + up to 8 companions)", 0, "annual", "Delta"),
            ("Zone 5 Priority Boarding", 0, "annual", None),
            ("20% Inflight Discount (food & beverages, excl. Wi-Fi)", 0, "annual", "Delta"),
            ("TakeOff 15 — 15% Off Award Flights", 0, "annual", None),
            ("$100 Delta Stays Credit (prepaid hotels/vacation rentals via delta.com/stays)", 100, "annual", "Hotels"),
            ("Uber One Credit (up to 6 months; promotional)", 9.99, "monthly", "Rideshare: Uber"),
        ], {
            # Base 1x on all purchases
            "_base": 1,
            # Delta purchases (direct): 2x total → 1 additional above base
            "Delta": 1,
            # Dining: 2x total → 1 additional above base
            "Dining": 1,
            # Groceries: 2x total → 1 additional above base
            "Groceries": 1,
        }),

        ("united_quest", "United Quest℠ Card", "United MileagePlus", "active", 350, [
            # Dollar credits
            ("United TravelBank Credit", 200, "annual", "United"),
            ("Renowned Hotels Credit", 150, "annual", "Renowned Hotels & Resorts"),
            ("JSX Flight Credit", 150, "annual", "Airlines"),
            ("Instacart Credit ($10+$5/mo, through 12/31/27)", 15, "monthly", None),
            ("Rideshare Credit (requires annual enrollment; $12 in December)", 8, "monthly", "Ground Transportation"),
            ("Avis/Budget Rental Credit (two $40 credits via United Cars)", 80, "annual", "Car Rental"),
            # Non-dollar perks
            ("Free Checked Bags (1st & 2nd bag, cardholder + 1 companion; ticket must be on card)", 0, "annual", "United"),
            ("10,000-Mile Anniversary Award Discount (auto-applied at anniversary)", 0, "annual", None),
            ("1,000 Card Bonus PQP", 0, "annual", None),
            ("Global Entry / TSA PreCheck / NEXUS Credit (every 4 years; using 'annual' as closest available frequency)", 120, "annual", None),
            ("United Inflight 25% Back (food, beverages, WiFi on United-operated flights)", 0, "annual", "United"),
            ("Priority Boarding", 0, "annual", None),
            ("Complimentary Instacart+ (3 months at opening, 50% off after through 12/31/27)", 0, "annual", None),
        ], {
            # multiplier = BONUS above base (total earn − 1).
            # Base 1x → multiplier=1 for base row.
            # United Purchases total 3x → additional 2x above base.
            # Renowned Hotels total 5x → additional 4x above base.
            # Travel (non-United) total 2x → additional 1x above base.
            # Dining total 2x → additional 1x above base.
            # Streaming total 2x → additional 1x above base.
            "_base": 1,
            "United Purchases": 2,
            "Renowned Hotels & Resorts": 4,
            "Airlines": 1,
            "Dining": 1,
            "Streaming": 1,
        }),

        # Business card — $0 annual fee; earns 2x Amex MR on all purchases up to $50K/year (then 1x).
        # Foreign transaction fee: 2.7%. No bonus categories. Welcome offer: 15K MR after $3K spend/3 months.
        ("amex_blue_business_plus", "The Blue Business® Plus Credit Card from American Express (Business)",
         "Amex MR", "not_held", 0, [
            # Amex Venue Collection: 10% back on concessions at participating stadiums/arenas,
            # up to $250 back per calendar year. Available to all US Amex cardholders.
            ("Amex Venue Collection Concessions Credit (10% back, up to $250/yr)", 250, "annual", None,
             "10% back on food & beverage concessions at Amex Venue Collection stadiums/arenas. Max $250/yr."),
            ("Extended Warranty Protection", 0, "annual", None,
             "Extends manufacturer warranty by up to 1 additional year on warranties of 5 years or less."),
            ("Purchase Protection", 0, "annual", None,
             "Covers eligible purchases against theft/accidental damage for 90 days. Up to $1,000/occurrence, $50,000/year."),
            ("Employee Cards at No Additional Cost", 0, "annual", None,
             "Add employee cards free; all spend earns 2x MR and counts toward the $50K annual cap."),
        ], {
            # Flat 2x Membership Rewards on all eligible purchases up to $50,000/calendar year.
            # Drops to 1x above $50K. No bonus categories — base rate IS the 2x rate.
            "_base": 2,
        }),

        ("atmos_ascent", "Atmos Rewards Ascent Visa Signature",
         "Atmos Rewards", "active", 95, [
            ("Free First Checked Bag", 0, "annual", None,
             "Cardholder + up to 6 companions on same reservation; Alaska/Hawaiian flights; must pay with card"),
            ("Preferred Boarding", 0, "annual", None,
             "Cardholder + up to 6 companions on same reservation"),
            ("Inflight Purchase Rebate", 0, "annual", None,
             "20% back on food, beverages, Wi-Fi on Alaska/Hawaiian flights"),
            ("Alaska Lounge+ Discount", 100, "annual", None,
             "Membership reduced from $795 to $695"),
        ], {
            "_base": 1,
            "Alaska Airlines": 2,    # total 3x → additional 2x above base
            "Hawaiian Airlines": 2,  # total 3x → additional 2x above base
            "Gas Stations": 1,       # total 2x
            "EV Charging": 1,        # total 2x
            "Cable": 1,              # total 2x
            "Streaming": 1,          # total 2x
            "Ground Transportation": 1,  # total 2x (Transit / Rideshare)
        }),

        ("united_explorer", "United℠ Explorer Card",
         "United MileagePlus", "active", 150, [
            ("Free First Checked Bag", 0, "annual", None,
             "Cardholder + 1 companion; United-operated flights; must pay with card"),
            ("Priority Boarding", 0, "annual", None, None),
            ("United Club Passes", 0, "annual", None,
             "2 one-time passes per year; no guest access"),
            ("Global Entry / TSA PreCheck / NEXUS", 120, "annual", None,
             "Every 4 years in practice — statement credit on application fee"),
            ("25% Back on United Inflight", 0, "annual", None,
             "Food, beverages, Wi-Fi on United-operated flights"),
            ("United Hotels Credit", 100, "annual", "Hotels",
             "Up to $50/stay, max 2 stays/year; prepaid via United Hotels portal"),
            ("Rideshare Credit", 5, "monthly", "Ground Transportation",
             "Requires annual enrollment"),
            ("Avis/Budget Rental Credit", 50, "annual", "Car Rental",
             "Up to $25/rental, max 2 rentals/year via cars.united.com"),
            ("JSX Flight Credit", 100, "annual", "Airlines", None),
            ("Instacart Credit", 10, "monthly", None,
             "Through 12/31/2027"),
        ], {
            "_base": 1,
            "United Purchases": 1,  # total 2x
            "Dining": 1,            # total 2x
            "Hotels": 1,            # total 2x (direct hotel purchases)
        }),

        # ── Pure Cash Back Cards ──────────────────────────────────────

        ("wells_fargo_active_cash", "Wells Fargo Active Cash", "Cash Back", "not_held", 0, [],
         {"_base": 2}),

        ("us_bank_cash_plus", "US Bank Cash+", "Cash Back", "not_held", 0, [],
         {"_base": 1}),
        # Note: 5% on 2 user-chosen categories (cap $2K/qtr), 2% on 1 user-chosen category.
        # User should set their chosen categories via CSC overrides on their card.

        ("bofa_customized_cash", "Bank of America Customized Cash Rewards", "Cash Back", "not_held", 0, [],
         {"_base": 1, "Groceries": 1}),
        # 3% on 1 user-chosen category (cap $2.5K/qtr), 2% grocery/wholesale, 1% else.
        # BofA Preferred Rewards members get 25-75% boost on all earn rates.

        ("amex_blue_cash_everyday", "Blue Cash Everyday", "Cash Back", "not_held", 0, [],
         {"_base": 1, "Groceries": 2, "Gas Stations": 2, "Online Shopping": 2}),
        # 3% grocery (cap $6K/yr then 1%), 3% gas, 3% online retail, 1% else.

        ("capital_one_savor", "Capital One Savor", "Cash Back", "not_held", 95, [],
         {"_base": 1, "Dining": 3, "Groceries": 2, "Streaming": 3, "Live Entertainment": 3}),
        # 4% dining/entertainment/streaming, 3% grocery, 1% else.

        ("citi_custom_cash", "Citi Custom Cash", "Citi ThankYou", "not_held", 0, [],
         {"_base": 1, "_auto_top": [
            "Dining", "Groceries", "Gas Stations", "Travel", "Transit",
            "Streaming", "Drugstore", "Home Improvement", "Fitness & Gyms", "Live Entertainment",
         ]}),
        # 5% on top eligible spend category each billing cycle (cap $500), 1% else.
        # Earns ThankYou Points, so it's in the Citi ThankYou ecosystem.

        ("target_redcard", "Target REDcard Credit", "Target Circle", "not_held", 0, [],
         {"_base": 1}),
        # 5% off Target purchases (modeled as 1x since it's a discount, not points).

        ("amazon_prime_visa", "Amazon Prime Visa", "Cash Back", "not_held", 0, [],
         {"_base": 1, "Amazon": 4, "Whole Foods": 4, "Dining": 1, "Gas Stations": 1, "Transit": 1}),
        # 5% Amazon/Whole Foods, 2% dining/gas/transit, 1% else.

        ("walmart_rewards_card", "Capital One Walmart Rewards", "Walmart Rewards", "not_held", 0, [],
         {"_base": 1, "Walmart": 4}),
        # 5% Walmart.com, 2% Walmart stores/restaurants/travel, 1% else.
    ]

    for product_key, card_name, eco_name, status, annual_fee, benefits, rates in products:
        eco_id = eco_map.get(eco_name)

        # Upsert product row — create if missing, update name/ecosystem if it drifted
        product = session.query(CardProduct).filter_by(product_key=product_key).first()
        if not product:
            product = CardProduct(product_key=product_key, status=status)
            session.add(product)
        product.card_name = card_name
        product.ecosystem_id = eco_id
        if annual_fee and not product.notes:
            product.notes = f"Annual fee: ${annual_fee}"
        session.flush()

        # Always refresh earning rates from seed (wipes stale data, applies fixes)
        session.query(CardProductReward).filter_by(product_id=product.id).delete()
        base_rate = rates.get('_base', 1)
        session.add(CardProductReward(
            product_id=product.id, points_category_id=None,
            multiplier=base_rate, is_base_rate=True, reward_type='fixed',
        ))
        # _auto_top: list of category names that compete for 5x each billing cycle
        auto_top_cats = rates.get('_auto_top', [])
        for cat_name in auto_top_cats:
            cat_id = cat_map.get(cat_name)
            if cat_id:
                session.add(CardProductReward(
                    product_id=product.id, points_category_id=cat_id,
                    multiplier=4, is_base_rate=False, reward_type='auto_top_category',
                ))
        for cat_name, additional in rates.items():
            if cat_name in ('_base', '_auto_top'):
                continue
            cat_id = cat_map.get(cat_name)
            if cat_id and additional > 0:
                session.add(CardProductReward(
                    product_id=product.id, points_category_id=cat_id,
                    multiplier=additional, is_base_rate=False, reward_type='fixed',
                ))

        # Refresh benefits only if the seed defines any (preserves manually-added ones
        # when benefits list is empty — use explicit None sentinel if you want to clear).
        # Upsert-by-name rather than delete-and-recreate: this function runs on every
        # startup, and CardBenefit.id is referenced by BenefitUsage (cascade-deletes on
        # remove) and by the frontend — deleting-and-recreating wiped usage history and
        # handed out new ids every restart.
        if benefits:
            existing_by_name = {
                b.benefit_name: b for b in
                session.query(CardBenefit).filter_by(product_id=product.id).all()
            }
            for ben_tuple in benefits:
                ben_name, amount, frequency, trigger = ben_tuple[:4]
                ben_notes = ben_tuple[4] if len(ben_tuple) > 4 else None
                existing = existing_by_name.get(ben_name)
                if existing:
                    existing.amount = amount
                    existing.reset_frequency = frequency
                    existing.trigger_category = trigger
                    existing.notes = ben_notes
                else:
                    session.add(CardBenefit(
                        product_id=product.id,
                        benefit_name=ben_name,
                        amount=amount,
                        reset_frequency=frequency,
                        trigger_category=trigger,
                        notes=ben_notes,
                        tracking_type=_infer_tracking_type(ben_name, frequency),
                    ))

    session.commit()
    print(f"Card products seeded: {session.query(CardProduct).count()} products in catalog")


def import_points_from_excel(filepath, session):
    """
    Import points data from the points Excel file into the proper schema:
      - 'valuations' sheet → PointsEcosystem table
      - 'points' sheet    → CardProduct + CardProductReward tables
                             + links Card.product_id

    Earning rate logic: each category column holds the ADDITIONAL points above Base.
    Total earn for a category = Base + category value.
    A value of 0 means no bonus (total = Base). None means category not applicable.

    The card_db_id column contains comma-separated Card.id values
    (e.g., "3,19" = same product held by Omer and Daniella).
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True)
    imported_ecos = 0
    imported_products = 0

    # ── Sheet 1: Valuations → PointsEcosystem ────────────────────────────
    if 'valuations' in wb.sheetnames:
        ws = wb['valuations']
        headers = [c.value for c in ws[1]]
        cash_types = {'Cash', 'cash'}
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(headers, row))
            name = str(d.get('ecosystem', '')).strip()
            if not name:
                continue
            eco_type = str(d.get('type', '')).strip()
            cons = d.get('conservative_cpp')
            yours = d.get('your_valuation_cpp')
            cons_basis = d.get('conservative_basis')
            existing = session.query(PointsEcosystem).filter_by(name=name).first()
            if existing:
                if cons is not None:
                    existing.conservative_cpp = float(cons)
                if yours is not None:
                    existing.your_cpp = float(yours)
                existing.is_cash_back = eco_type in cash_types
                existing.eco_type = eco_type
                if cons_basis:
                    existing.conservative_basis = str(cons_basis)
            else:
                session.add(PointsEcosystem(
                    name=name,
                    currency_name=name,
                    eco_type=eco_type,
                    conservative_cpp=float(cons) if cons else 1.0,
                    your_cpp=float(yours) if yours else float(cons) if cons else 1.0,
                    is_cash_back=eco_type in cash_types,
                    conservative_basis=str(cons_basis) if cons_basis else None,
                ))
                imported_ecos += 1
        session.flush()

    # ── Sheet 2: Points → CardProduct + CardProductReward ─────────────────
    if 'points' in wb.sheetnames:
        ws = wb['points']
        headers = [c.value for c in ws[1]]

        # Category columns = everything after the metadata columns
        meta_cols = {'product_key', 'card_db_id', 'ecosystem', 'status', 'card_name', 'notes', 'Base'}
        cat_columns = [h for h in headers if h and h not in meta_cols]

        # Ensure PointsCategory rows exist for every category column
        cat_map = {}
        for col_name in cat_columns:
            cat = session.query(PointsCategory).filter_by(name=col_name).first()
            if not cat:
                max_order = session.query(PointsCategory).count() + 1
                cat = PointsCategory(name=col_name, display_order=max_order)
                session.add(cat)
                session.flush()
            cat_map[col_name] = cat.id

        eco_map = {e.name: e.id for e in session.query(PointsEcosystem).all()}

        for row in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(headers, row))
            product_key = str(d.get('product_key', '')).strip()
            if not product_key:
                continue
            card_name = str(d.get('card_name', '')).strip() or product_key
            eco_name = str(d.get('ecosystem', '')).strip()
            status = str(d.get('status', 'active')).strip()
            notes = d.get('notes')
            base_rate = d.get('Base')
            db_ids_str = str(d.get('card_db_id', '')).strip()

            eco_id = eco_map.get(eco_name)

            # ── Upsert CardProduct ────────────────────────────────────────
            product = session.query(CardProduct).filter_by(product_key=product_key).first()
            if not product:
                product = CardProduct(
                    product_key=product_key,
                    card_name=card_name,
                    ecosystem_id=eco_id,
                    status=status,
                    notes=str(notes) if notes else None,
                )
                session.add(product)
                session.flush()
            else:
                product.card_name = card_name
                product.ecosystem_id = eco_id
                product.status = status
                if notes:
                    product.notes = str(notes)

            # ── Replace rewards (delete + re-insert) ──────────────────────
            session.query(CardProductReward).filter_by(product_id=product.id).delete(
                synchronize_session=False)

            if base_rate is not None:
                # Base rate
                session.add(CardProductReward(
                    product_id=product.id, points_category_id=None,
                    multiplier=float(base_rate), is_base_rate=True,
                ))
                # Category bonus rates (additional above base)
                for col_name in cat_columns:
                    val = d.get(col_name)
                    if val is not None and isinstance(val, (int, float)) and val > 0:
                        cat_id = cat_map.get(col_name)
                        if cat_id:
                            session.add(CardProductReward(
                                product_id=product.id, points_category_id=cat_id,
                                multiplier=float(val), is_base_rate=False,
                            ))

            # ── Link Card → CardProduct ───────────────────────────────────
            if db_ids_str and not db_ids_str.startswith('⚠'):
                try:
                    card_db_ids = [int(x.strip()) for x in db_ids_str.split(',') if x.strip().isdigit()]
                except ValueError:
                    card_db_ids = []
                for card_db_id in card_db_ids:
                    card = session.query(Card).filter_by(id=card_db_id).first()
                    if card:
                        card.product_id = product.id
                        # Also set ecosystem on card for backward compat
                        if eco_id:
                            card.ecosystem_id = eco_id

            imported_products += 1

    session.commit()
    return {'ecosystems_imported': imported_ecos, 'products_imported': imported_products}


def import_cards_from_excel(filepath, session):
    """Import cards from Excel file"""
    import openpyxl
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    imported = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        d = dict(zip(headers, row))
        card_id = str(d.get('ID', '')).strip()
        if not card_id:
            continue
        existing = session.query(Card).filter_by(card_id=card_id).first()
        if not existing:
            # Parse close_date if present
            close_date = d.get('Close Date')
            if close_date and hasattr(close_date, 'year'):
                pass  # already a datetime
            else:
                close_date = None

            card = Card(
                card_id=card_id,
                last_four=d.get('Last 4'),
                issuer=str(d.get('Issuer', '')).strip() or None,
                brand=str(d.get('Brand', '')).strip() or None,
                card_name=str(d.get('Card', '')).strip() or None,
                network=str(d.get('Network', '')).strip() or None,
                issue_date=d.get('Issue Date'),
                close_date=close_date,
                annual_fee=d.get('Annual Fee'),
                credit_limit=d.get('Credit Limit'),
                is_active=close_date is None,
            )
            session.add(card)
            imported += 1
    session.commit()
    return imported

